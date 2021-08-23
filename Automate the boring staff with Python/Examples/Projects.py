## projects from Automate the boring stuff with Python
## Erim Serdönmez

## -------------------------------- Chapter 3 Functions -----------------------------------##

## The Collats Sequence
    #   Write a function named collatz() that has one parameter named number. If number is even, then collatz() should print number // 2 and return this value. If number is odd, then collatz() should print and return 3 * number + 1.

'''def collatz(number):
    number = int(number)
    if (number % 2) == 0:
        print(number // 2)
    else:
        print(3 * number +1)
'''

## Input Validation
    #   Add try and except statements to the previous project to detect whether the user types in a noninteger string. Normally, the int() function will raise a ValueError error if it is passed a noninteger string, as in int('puppy'). In the except clause, print a message to the user saying they must enter an integer.

'''print('Hello and Welcome\n ...Please enter a number...\n press "q" to quit.') 

while True:
 
    try:
        prompt = input()
        if prompt == 'q':
            break
        collatz(prompt)
    except ValueError:
        print('\n You must enter a number and not a string!!!')'''

## -------------------------------- Chapter 4 Lists ---------------------------------------##

# Comma Code 

    # Write a function that takes a list value as an argument and returns a string with all the items separated by a comma and a space, with and inserted before the last item. For example, passing the previous spam list to the function would return 'apples, bananas, tofu, and cats'. But your function should be able to work with any list value passed to it. Be sure to test the case where an empty list [] is passed to your function.

'''def comma(list):
    if len(list) == 0:
        print('This is an empty list.')
    else:
        list.insert(-1,'and')
        string = ','.join(list)
        print(string)
        
    
    
spam = ['apples','bananas','tofu','cats']
bacon = []
comma(spam)
comma(bacon)'''

# Coin Flip Streaks

'''import random

numberOfStreaks = 0
coinFlip = []
streak = 0


for experimentNumber in range(10000):
    # code that creates a list of 100 'heads' or 'tails' values
    for i in range(100):
        coinFlip.append(random.randint(0,1))

# code that checks if there is a streak of 6 head or tails in a row

for i in range(len(coinFlip)):
    if i == 0:
        pass    
    elif coinFlip[i] == coinFlip[i-1]: # #checks if current list item is the same as before
        streak += 1
    else:
        streak = 0

    if streak == 6:
        numberOfStreaks += 1

print("Chance of streak : %s%%" % (numberOfStreaks / (100*10000)))'''

# Character picture Grid

'''grid = [['.', '.', '.', '.', '.', '.'],
        ['.', 'O', 'O', '.', '.', '.'],
        ['O', 'O', 'O', 'O', '.', '.'],
        ['O', 'O', 'O', 'O', 'O', '.'],
        ['.', 'O', 'O', 'O', 'O', 'O'],
        ['O', 'O', 'O', 'O', 'O', '.'],
        ['O', 'O', 'O', 'O', '.', '.'],
        ['.', 'O', 'O', '.', '.', '.'],
        ['.', '.', '.', '.', '.', '.']]
'''


## -------------------------------- Chapter 11 --------------------------------------------##

# Debugging the coin Toss.
# theres a bug in this code find it.


'''import random,logging
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)
logging.debug('Star of the program')
guess = ''
score = 0
for i in range(0,5):
    logging.info('Start of the for loop'+ str(i))
    print('Guess the coin toss! Enter heads or tails:')
    guess = input()         # first of all this has to be an int for comperison
    toss = random.randint(0, 1) # 0 is tails, 1 is heads
    if toss == 0:
        toss = 'heads'
        if toss.lower() == guess.lower():
            score += 1
            print('You got it!')
            
        else:
            print('Sory bad luck.')
    elif toss == 1:
        toss = 'tails'
        if toss.lower() == guess.lower():
            score += 1
            print('you got it')
            
        else:
            print('Sory bad luck')
    if i == 4:
        print(f'Thanks for playing this is your score: {score}')'''
'''else:    
    print('Nope! Guess again!')
    guess = input()       # this guesss variable is not the sameas guess, also again not integer
    if toss == guess:
        print('You got it!')
    else:
        print('Nope. You are really bad at this game.')'''

## --------------------------------- CHAPTER 12 -----------------------------------------##

# Command Line Emailer #

# Write a program that takes email address and string of text on the command line
# Using Selenium logs in to your email account 
# Send an email of the string to the provided adress.

'''
import sys
if len(sys.argv) > 2:
    testEmail = sys.argv[1]
    testString = ' '.join(sys.argv[2:])
else:
    print('Moving On for right now.')

#print(testEmail,testString)

## Using Selenium logs in to your email account 

from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys

browser = webdriver.Firefox()
browser.get('https://mail.yandex.com.tr/?noretpath=1')
time.sleep(3)
firstElement = browser.find_element_by_css_selector('.button2_theme_mail-white')
firstElement.click()
time.sleep(3)
emailElem = browser.find_element_by_name('login')
emailElem.send_keys('erimserdnmez@yandex.com')
emailElem.submit()
time.sleep(3)
passwordElem = browser.find_element_by_name('passwd')
passwordElem.send_keys('92aF35Gdr10e8'), passwordElem.submit()
time.sleep(3)
try:
    skipElement = browser.find_element_by_css_selector('.Button2_view_pseudo')
    skipElement.click()
except Exception:
    time.sleep(3)
    nextElement = browser.find_element_by_css_selector('.mail-ComposeButton')
    nextElement.click()
    time.sleep(3)
    recieverElem = browser.find_element_by_css_selector('.ComposeRecipients-ToField > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1)')
    recieverElem.send_keys(testEmail)
    
    topicElement = browser.find_element_by_css_selector('.composeTextField')
    topicElement.click()
    topicElement.send_keys('Test')
    bodyElement = browser.find_element_by_css_selector('.cke_wysiwyg_div')
    bodyElement.click()
    bodyElement.send_keys(testString)
    time.sleep(3)
    sendElement = browser.find_element_by_css_selector('.ComposeControlPanel-SendButton > button:nth-child(1)')
    sendElement.click()
'''
# Image Site Downloader #

# Write a program that goes to Imgur
# Searches for a category of photos
# downloads all the resulting images.
'''
import requests,os,bs4,time
from selenium import webdriver

def downloader(query,max_save,output_path):
    
    Args:
         (str): search query
        max_save (int) : max number of images to save to results
    Returns:
        None
    '''
    # Create imgur search url
'''
    searchUrl = 'https://imgur.com/search'
    queryUrl = searchUrl + '?q' + query

    # set up output_path
    abs_output_path = os.path.abspath(output_path)
    os.makedirs(abs_output_path,exist_ok=True)

    # Make request to imgur with querry
    res1 = requests.get(queryUrl)

    try:
        res1.raise_for_status()

        # Parse res.text with bs4 to images
        imugurSoup = bs4.BeautifulSoup(res1.text,'html.parser')
        images = imugurSoup.select('.image-list-link')

        # Extract number image urls
        num_to_save = min(max_save,len(images))
        download_links = ['https:' +img.get('src') for img in images[:num_to_save]]

        # Make requests for extracted url
        for link in download_links:

            # Request image link from imgur
            res2 = requests.get(link)

            try:
                res2.raise_for_status()

                # save to file with url base name in folder results
                imgFile = open(os.path.join(abs_output_path,os.path.basename(link)), 'wb')
                for chunk in res2.iter_content(100000):
                    imgFile.write(chunk)
                imgFile.close()
            except Exception as exc:
                print('There was a problem : %s' % (exc))

    except Exception as exc:
        print('There was a problem: %s' % (exc))

    

if __name__ == '__main__':
    downloader('metallica', 10, 'results')
'''

# Playing 2048 #
# Write a program which goes to 'https://gabrielecirculli.github.io/2048/'
# Play the game with Selenium keys module 

'''from selenium import webdriver
from selenium.webdriver.common.keys import Keys

browser = webdriver.Firefox()
browser.get('https://play2048.co')

htmlElem = browser.find_element_by_tag_name('html')
htmlElem.send_keys(Keys.UP)'''


## ------------------------------ Chapter 13 -----------------------------------------##

# Multiplication Table Maker

# Create a program that takes number N from the command line and creates and NxN multiplication table in excel spreadsheet.
'''
import sys,openpyxl
from openpyxl.cell import cell
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title ='Multiplication Table'
wb.save('multiplicationTable.xlsx')
#prompt =''.join(sys.argv[1:])
prompt = input()
prompt =''.join(prompt)


if prompt.isnumeric():
    for i in range(int(prompt)+1):
        for y in range(int(prompt)+1):
            isHeader = False

            if y == 0 and i == 0:
                isHeader = True
                n = ''
            
            elif i == 0:
                isHeader=True
                n = y
            
            elif y == 0:
                isHeader = True
                n = i

            else:
                n = i*y

            c = sheet.cell(row=y+1,column=i+1)

            if isHeader:
                c.font = Font(bold=True)

            c.value = n
    f= 'multiplication_table_' +str(prompt) + '.xlsx'

    wb.save(f)

    print('Saved as ' + f)
        
       
else:
    print('You need to enter a number')

'''

## ----------------------------------- Chapter 14 ---------------------------------------##

## Downloading Google Forms Data

'''import ezsheets

ss = ezsheets.Spreadsheet('1l88mn7zn6CrU4G7HqGU2TO5xC8h1WaOeEnNVy2vsByY')
sheet = ss[0]
print(sheet.columnCount)
sheet.columnCount = 5
sheet.rowCount = 5
email = sheet.getColumn(3)

ss2 = ezsheets.Spreadsheet('Email_Adresses')
for i,value in enumerate(email):
    email[i] = value

sheet1 = ss2[0]

sheet1.updateColumn(1,email)
sheet1.columnCount = 5
sheet1.rowCount = 5
sheet1.title = 'Email Adresses'

ss2.downloadAsPDF()'''

## Finding Mistakes in a Spreadsheet
'''
import ezsheets
'''
'''
ss = ezsheets.Spreadsheet('1jDZEdvSIh4TmZxccyy0ZXrH-ELlrwq8_YYiZrEOB4jg')
ss.downloadAsExcel('BeanCount.xlsx')
ezsheets.upload('BeanCount.xlsx')
'''
'''
ss = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/1yI4r0TzCF_rIejTPIim_Gd6EbdRIUb7bs7ZbUipy-L4/edit#gid=1531551185')

def CheckData(ss):

    sheet = ss[0]
    y = 0
    wrong_answers = []
    for i in range(len(sheet.getRows())):
        if i == (1 and 0):
            continue
        if i > 1:
            try:
                if int(sheet.getRow(i)[y]) * int(sheet.getRow(i)[y+1]) == int(sheet.getRow(i)[y+2]):
                    continue
                else:
                    answer = sheet.getRow(i)
                    wrong_answers.append(i)
                    wrong_answers.append(answer)
            except ValueError:
                continue
    print(wrong_answers)

sheet = ss[0]

sheet.updateRow(14399,['804','623','500892'])
sheet.refresh()

CheckData(ss)'''

## ------------------------------------ Chapter 15 ----------------------------------------##

## PDF Paranoia

# Using the os.walk() function write a script that will go through every PDF in a folder (and its subfolders) and encrypt the PDFs using a password provided on the command line.
# Save each encrpyted PDF with an _encrypted.pdf added to the original filename.
# Before deleting the file, have the program attempt to read and decrypt the file to ensure that it was encrypted correctly.

# Then, write a program that finds all encrypted PDFs in a folder(and its subfolders) and creates a decrpyted copy of the PDF using a password.
# If he password is incorrect, the program should print a message to the user and continue to the next PDF.

# Using the os.walk() function:

'''import os,PyPDF2
from PyPDF2.utils import PdfReadError

password = input('\n What will be the password for these PDF files?')'''

"""for foldername,subfolders,filenames in os.walk('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials'):              
    for filename in filenames:
        if filename.endswith('_encrypted.pdf'):
            os.unlink(foldername + '\\' + filename)
        if filename.endswith('_decrypted.pdf'):
            os.unlink(foldername+'\\'+filename)"""

'''for foldername,subfolders,filenames in os.walk('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials'):              
    for filename in filenames:
        if filename.endswith('.pdf'):
            try:
                encryptedReader = PyPDF2.PdfFileReader(foldername+'\\'+filename)
                if encryptedReader.isEncrypted == False:
                    pdfFileObj = open(foldername+ '\\'+filename,'rb')
                    pdfReaderObj = PyPDF2.PdfFileReader(pdfFileObj)
                    pdfWriterObj = PyPDF2.PdfFileWriter()
                    for pageNum in range(pdfReaderObj.numPages):
                        pageObj = pdfReaderObj.getPage(pageNum)
                        pdfWriterObj.addPage(pageObj)
                    outputPdf = open(foldername+ '\\' + filename[:-4] +'_encrypted.pdf','wb')
                    pdfWriterObj.encrypt(password)
                    pdfWriterObj.write(outputPdf)
                    print(f'This file has been en-crypted {filename}') 
                    print(f'Deleting the file \n {filename}')
                    pdfFileObj.close()
                    outputPdf.close()
                    os.unlink(foldername+'\\' + filename)
            except PdfReadError:
                print(f'This file is already encrypted {filename} moving on the de-cryption')          
            if encryptedReader.isEncrypted == True:
                if encryptedReader.decrypt(password):
                    encryptedWriter = PyPDF2.PdfFileWriter()
                    for pageNum in range(encryptedReader.numPages):
                        pageObj = encryptedReader.getPage(pageNum)
                        encryptedWriter.addPage(pageObj)
                    decrypted_file = open(foldername+'\\' + filename[:-14]+'_decrypted.pdf','wb')
                    encryptedWriter.write(decrypted_file)
                    print('This file has been de-crypted: ' + filename)
                    print(f'Deleting this file: {filename}')
                    decrypted_file.close()
                    os.unlink(foldername+'\\'+filename)
                else:
                    print(f'Password for {filename} is wrong.')'''
            
                

## ------------------------------------- Chapter 16 --------------------------------------##


# Excel-to-CSV Converter # 

'''
Excel can save a spreadsheet to a CSV file with a few mouse clicks, but if you had to convert hundreds of Excel files to CSVs, it would take hours of clicking. Using the openpyxl module from Chapter 12, write a program that reads all the Excel files in the current working directory and outputs them as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create one CSV file per sheet. The filenames of the CSV files should be <excel filename>_<sheet title>.csv, where <excel filename> is the filename of the Excel file without the file extension (for example, 'spam_data', not 'spam_data.xlsx') and <sheet title> is the string from the Worksheet object’s title variable.

This program will involve many nested for loops. The skeleton of the program will look something like this:


import csv,openpyxl,os


os.chdir('C:\\Users\\devil\\Desktop\\Project 16')

## For Deleting the files for retry:

for filename in os.listdir('.'):
    if filename.endswith('_Sheet.csv'):
        os.unlink(filename)
'''
        
'''
for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFile = open(excelFile[:-5] + '_' + sheetName+ '.csv','w',newline='')
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)

        
        # Loop through every row in the sheet.
        for rowNum in range(1,sheet.max_row+1):
            rowData =[]
            # Loop through each cell in the row.
            for colNum in range(1,sheet.max_column+1):
                # Append each cell's data to rowData
                rowData.append(sheet.cell(row=rowNum,column=colNum).value)
                # Write the rowData list to the CSV file.
                if colNum == 10:
                    csvWriter.writerow(rowData)
        csvFile.close()

'''

##---------------------------------- Chapter 17 --------------------------------------------##

## Prettified Stopwatch 

# Expand the stopwatch project from this chapter so that it uses the rjust() and ljust() string methods to “prettify” the output. (These methods were covered in Chapter 6.) Instead of output such as this:
'''
    Lap #1: 3.56 (3.56)
    Lap #2: 8.63 (5.07)
    Lap #3: 17.68 (9.05)
    Lap #4: 19.11 (1.43)
'''

# ... the output will look like this:
'''
    Lap # 1:   3.56 (  3.56)
    Lap # 2:   8.63 (  5.07)
    Lap # 3:  17.68 (  9.05)
    Lap # 4:  19.11 (  1.43)
'''

# Note that you will need string versions of the lapNum, lapTime, and totalTime integer and float variables in order to call the string methods on them.

# Next, use the pyperclip module introduced in Chapter 6 to copy the text output to the clipboard so the user can quickly paste the output to a text file or email.

#! python3
# stopwatch.py -- Simple stopwatch program

import time,pyperclip

# Display the program's instructions.

print('Press ENTER to begin. Afterward, press ENTER to "click" the stopwatch.\nPress Ctrl+C to quit.')

# Press ENTER to begin
input()

print('Started')

# Get the first lap's start time.
starTime = time.time()
lastTime = starTime

lapNum = 1

# TODO : Start tracking the lap times
try:
    while True:
        input()
        lapTime = round(time.time() - lastTime,2)
        totalTime = round(time.time() - starTime,2)
        output = pyperclip.copy("Lap #%s: %s(%s)"%(str(lapNum).ljust(5),str(totalTime).center(10),str(lapTime).rjust(1)))
        print('Lap #%s: %s(%s)'%(str(lapNum).ljust(5),str(totalTime).center(10),str(lapTime).rjust(1)),end='')
        lapNum += 1
        # Reset the last lap time
        lastTime = time.time()
except KeyboardInterrupt:
    # Handle the Ctrl+C exception to keep its error message from displaying.
    print('\nDone...')

# Yeni branche atıldı

