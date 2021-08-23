#! python3
# updatingSpreadsheet.py --  Corrects costs in produceSales Spreadsheet.

## This program will do:
    # 1.Loops over all the rows
    # 2.If the row is for garlic,celery or lemons, changes the price

## Code will do the following:
    # 1.Open the spreadsheet File.
    # 2.For each row, check whether the value in column A is Celery,Garlic or Lemon.
    # 3.If it is, update the price in column B.
    # 4.Save the spreadsheet to a new file ( so that we dont lose the old spreadsheet)

## Step 1: Set up a Data Structure with he Update information.
'''
The Prices that We need to update as follows:
    Celery          1.19
    Garlic          3.07
    Lemon           1.27

Essentially what we will write is:
if produceName == 'Celery':
    cellObj = 1.19
if produceName == 'Garlic':
    cellObj = 3.07
if produceName == 'Lemon':
    cellObj = 1.27
'''

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices.
PRICE_UPDATES = {'Garlic':3.07,'Celery':1.19,'Lemon':1.27}
# If you need to update the Sheet again you only need to update above dictionary and not any other code

## Step 2: Check all Rows and update Incorrect Prices


#  Loop through the rows and update the prices.

for rowNum in range(2,sheet.max_row):       # Skip the first row
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum,column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
