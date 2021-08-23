## ----------------------This program says hello and ask for my name-----------------##

'''print("hello World")
print("What is your name? ")# ask for their name
myName = input()
print("It is good to meet you, " + myName)
print("The lenght of your name is : ")
print(len(myName))
print("What is your age? ")# ask for their age
myAge = input()
print("You will be " + str(int(myAge)+1 ) + " in a year.")'''

'''while True:
    print("Who are you?")
    name = input()
    if name != "Joe":
        continue
    print("Hello Joe What is your password?")

    password = input()

    if password == "swordfish":
        print("Access granted.")
        break'''

##---------------------------Truthy and Falsey ------------------------##

'''name = ""

while not name:
    print("Enter your name:")
    name = input()
print("How many guests will you have? ")
numOfGuests = int(input())
if numOfGuests:
    print("Be sure to have enough room for all your guests...")
print("Done...")'''

#--------------------------Calculating the sum of number from 0 to 100 ------------------##
'''
total = 0

for i in range(101):
    total = total + i
print(total)'''

'''import random

for i in range(5):
    print(random.randint(1,10))

import sys

while True:
    print("Type exit to exit.")
    response = input()
    if response == "exit":
        sys.exit()
    print("You typed " + response + " .")'''

## Guess the number game

'''import random

answer = int(random.randint(1,21))

print("I am thinking of a number between 1 and 20 ")

for i in range(1,7):
    response = int(input("\nTake a Guess..."))
    if response < answer:
        print("Your guess is too low.")
    elif response > answer:
        print("Your guess is too high")
    elif answer == response:
        print("\nGreat job you guessed it Correct!!!")
        print("it took you " + str(i) + " times to guess it.")
        break    
    if i == 6:
        print("Sorry you are out of chances.")
        print("The number I chosed was " + str(answer))'''


## Rock, Paper and Scissors my take
'''import random , sys'''

'''Rock = 0
Paper = 0
Scissors = 0
draw = 0
r = 1
p = 2
s = 3
aiNumber = int(random.randint(1,3))
#print(aiNumber)
print("Welcome to the rock , papar and scissors game")

x = input("\n Press any key to continue")
if x != " ":
    for i in range(10):
        aiNumber = int(random.randint(1,3))
        game = input("Press the indicated keys to play. \n (r)ock,(p)aper, (s)cissors \n Press 'q' to quit the game.")
        if game == "r":
            print("\nYou have selected Rock!!")
            if aiNumber == p:
                print("Sorry, the ai has selected Paper \n You lose this round.")
                Paper = Paper + 1
            if aiNumber == s :
                print("Gratz!! ai has selected Scissors, \n you have won this round.")
                Rock = Rock + 1 
            if aiNumber == r :
                print("Its a draw...")
                draw = draw + 1 
        if game == "s":
            print("\nYou have selected Scissors!!")
            if aiNumber == r:
                print("Sorry, the ai has selected Rock \n You lose this round.")
                Rock = Rock + 1
            if aiNumber == p:
                print("Gratz!! ai has selected Paper, \n you have won this round.")
                Scissors = Scissors + 1 
            if aiNumber == s:
                print("Its a draw...")
                draw = draw + 1 
        if game == "p":
            print("\nYou have selected Paper!!")
            if aiNumber == s:
                print("Sorry, the ai has selected Scissors \n You lose this round.")
                Scissors = Scissors + 1
            if aiNumber == r:
                print("Gratz!! ai has selected Rock, \n you have won this round.")
                Paper = Paper + 1 
            if aiNumber == p:
                print("Its a draw...")
                draw = draw + 1 
        if game == "q":
            print("Thank you for playing overall score is : \n")
            print("Rock: " + str(Rock) + " Paper: " + str(Paper) + " Scissors: " + str(Scissors) + " In total draw number was : " + str(draw))            
            sys.exit
            break      
        if i == 9:
            print("Thank you for playing overall score is : \n")
            print("Rock: " + str(Rock) + " Paper: " + str(Paper) + " Scissors: " + str(Scissors) + " In total draw number was : " + str(draw))  
            break'''
            

### Rock paper and scissors lesson version

'''print('ROCK, PAPER, SCISSORS')

# These variables keep track of the number of wins, losses, and ties.
wins = 0
losses = 0
ties = 0

while True: # The main game loop.
    print('%s Wins, %s Losses, %s Ties' % (wins, losses, ties))
    while True: # The player input loop.
        print('Enter your move: (r)ock (p)aper (s)cissors or (q)uit')
        playerMove = input()
        if playerMove == 'q':
            sys.exit() # Quit the program.
        if playerMove == 'r' or playerMove == 'p' or playerMove == 's':
            break # Break out of the player input loop.
        print('Type one of r, p, s, or q.')

    # Display what the player chose:
    if playerMove == 'r':
        print('ROCK versus...')
    elif playerMove == 'p':
        print('PAPER versus...')
    elif playerMove == 's':
        print('SCISSORS versus...')

    # Display what the computer chose:
    randomNumber = random.randint(1, 3)
    if randomNumber == 1:
        computerMove = 'r'
        print('ROCK')
    elif randomNumber == 2:
        computerMove = 'p'
        print('PAPER')
    elif randomNumber == 3:
        computerMove = 's'
        print('SCISSORS')

    # Display and record the win/loss/tie:
    if playerMove == computerMove:
        print('It is a tie!')
        ties = ties + 1
    elif playerMove == 'r' and computerMove == 's':
        print('You win!')
        wins = wins + 1
    elif playerMove == 'p' and computerMove == 'r':
        print('You win!')
        wins = wins + 1
    elif playerMove == 's' and computerMove == 'p':
        print('You win!')
        wins = wins + 1
    elif playerMove == 'r' and computerMove == 'p':
        print('You lose!')
        losses = losses + 1
    elif playerMove == 'p' and computerMove == 's':
        print('You lose!')
        losses = losses + 1
    elif playerMove == 's' and computerMove == 'r':
        print('You lose!')
        losses = losses + 1'''

##------------------------------ Functions ----------------------------------------------##

'''import random

def getAnswer(answerNumber):
    if answerNumber == 1:
        return "It is certain"
    elif answerNumber == 2:
        return "It is decidedly so"
    elif answerNumber == 3:
        return "Yes"
    elif answerNumber == 4:
        return "Reply hazy try again"
    elif answerNumber== 5:
        return "Ask again later"
    elif answerNumber == 6:
        return "Concentrate and ask again"
    elif answerNumber == 7:
        return "My reply is no"
    elif answerNumber == 8:
        return "Outlook not so good"
    elif answerNumber == 9:
        return "Very doubtful"

print(getAnswer(random.randint(1,9))) # You can use expressions in function calls'''

## Call Stacks

'''def a():
    print("a() starts")
    b()
    d()
    print("a() returns")

def b():
    print(" b() starts")
    c()
    print("b() returns")

def c():
    print("c() starts")
    print("c() returns")

def d():
    print(" d() starts")
    print("d() returns")

a()'''

# Both are local variables
'''
def spam():
    eggs = 99
    bacon()
    print(eggs)

def bacon():
    ham = 101
    eggs = 0

spam()'''

'''def spam():
    global eggs
    eggs = "spam" # this is the global
def bacon():
    eggs = "bacon" # this is local
def ham():
    print(eggs) # this is global
eggs = 42 # this is global
spam()
print(eggs) # global variable has changed by a local variable'''


##---------------------------------- try and except blocks ------------------------------##

'''def spam(divideBy):
    return 42 / divideBy

try:
    print(spam(2))
    print(spam(12))
    print(spam(0))
    print(spam(1))
except ZeroDivisionError: # will not return to try block but rather will continue down as normal.
    print('Error: Invalid argument.')'''

## Zigzag program##

'''import time,sys
indent = 0 # How many spaces to indent.
indentIncreasing = True # Whether the indentation is increasing or not

try:
    while True: # Main program loop.
        print(" " * indent, end="")
        print("**********")
        time.sleep(0.1) # Pause for 1/10 of a second

        if indentIncreasing:
            # Increase the number of spaces:
            indent = indent + 1 
            if indent == 20:
                #Change direction
                indentIncreasing = False
        else:
            #Decrease the number of spaces
            indent = indent - 1
            if indent == 0:
                # Change direction
                indentIncreasing = True
except KeyboardInterrupt:
    sys.exit()'''

## Collatz Sequence

'''def collatz(number):
    if number % 2 == 0:
        print(number//2)
    elif number %2 == 1:
        print(3* number+1)
while True:
    try:
        answer = input("Please enter a number: ")
        collatz(int(answer))
    except ValueError:
        print("Please write a number.")'''

## Lists 

'''catNames = []

while True:
    print("Enter the name of the cat " + str(len(catNames)+1) + " or enter nothing to stop.")
    name = input()
    if name == "":
        break
    catNames = catNames + [name] # list concatenation

print("Cat names are: ")
for name in catNames:
    print(" " + name)'''

'''supplies = ["pens", "staplers","flamethrowers","binders"]
for i in range(len(supplies)):
    print("Index " + str(i)+ " in supplies is: " + supplies[i])'''

## ------- Conway's game of Life ------ ##
'''import random,time,copy
WIDTH = 60
HEIGHT = 20

# Create a listof list for the cells:

nextCells = []

for x in range(WIDTH):
    column = [] # Create a new column
    for y in range(HEIGHT):
        if random.randint(0,1) == 0:
            column.append("#") # Add a living cell.
        else:
            column.append(" ") # Add a dead cell.

    nextCells.append(column) # NextCells is a list of column lists.

while True: # Main program loop
    print("\n\n\n\n\n") # Separate each step with newlines.
    currentCells = copy.deepcopy(nextCells)
    ## Print currentCells on the screen:
    for y in range(HEIGHT):
        for x in range(WIDTH):
            print(currentCells[x][y], end="") # Print the # or space
        print() #print a newline at the end of the row
        # Calculate the next step's cells based on current step's cells:
    for x in range(WIDTH):
        for y in range(HEIGHT):
            # Get neighboring coordinates:
            # "% WIDHT" ensures leftCoord is always between 0 and WIDTH -1
            leftCoord = (x-1)%WIDTH
            rightCoord = (x+1)%WIDTH
            aboveCoord = (y-1)% HEIGHT
            belowCoord = (y+1)%HEIGHT

            # Count number of living neighbors:
            numNeighbors = 0
            if currentCells[leftCoord][aboveCoord] == '#':
                numNeighbors += 1 # Top-left neighbor is alive.
            if currentCells[x][aboveCoord] == '#':
                numNeighbors += 1 # Top neighbor is alive.
            if currentCells[rightCoord][aboveCoord] == '#':
                numNeighbors += 1 # Top-right neighbor is alive.
            if currentCells[leftCoord][y] == '#':
                numNeighbors += 1 # Left neighbor is alive.
            if currentCells[rightCoord][y] == '#':
                numNeighbors += 1 # Right neighbor is alive.
            if currentCells[leftCoord][belowCoord] == '#':
                numNeighbors += 1 # Bottom-left neighbor is alive.
            if currentCells[x][belowCoord] == '#':
                numNeighbors += 1 # Bottom neighbor is alive.
            if currentCells[rightCoord][belowCoord] == '#':
                numNeighbors += 1 # Bottom-right neighbor is alive.

                # Set cell based on Conway's Game of Life rules:
            if currentCells[x][y] == '#' and (numNeighbors == 2 or numNeighbors == 3):
                # Living cells with 2 or 3 neighbors stay alive:
                nextCells[x][y] = '#'
            elif currentCells[x][y] == ' ' and numNeighbors == 3:
                # Dead cells with 3 neighbors become alive:
                nextCells[x][y] = '#'
            else:
                # Everything else dies or stays dead:
                nextCells[x][y] = ' '
    time.sleep(1) # Add a 1-second pause to reduce flickering.'''

## Coin Flip Streaks ##

'''import random

numberOfStreaks = 0
coinFlip = []
streak = 0


for experimentNumber in range(10000):
    # code that creates a list of 100 'heads' or 'tails' values
    for i in range(100):
        coinFlip.append(random.randint(0,1))

# code that checks if there is a streak of 6 head or tails in a row

for i in range(len(coinFlip)):
    if i == 0:
        pass    
    elif coinFlip[i] == coinFlip[i-1]: # #checks if current list item is the same as before
        streak += 1
    else:
        streak = 0

    if streak == 6:
        numberOfStreaks += 1

print("Chance of streak : %s%%" % (numberOfStreaks / (100*10000)))'''

## ---------------------------- Dictionaries-------------------------------------------- ##

'''birthday = {"Alice":"Apr 1","Bob":"Dec 12","Carol":"Mar 4"}

while True:
    print("Enter a name : (blank to quit)")
    name = input()
    if name == "":
        break
    if name in birthday:
        print(birthday[name] + " is the birthday of " + name)
    else:
        print("I do not have birthday information for " + name)
        print("What is their birthday? ")
        bday = input()
        birthday[name] = bday
        print("Birthday data base updated.")'''


## Program loops over each character in the message then counts the characters an puts them in a dict.

'''import pprint # for better sorting and display


message = "It was a bright cold day in April, and the clocks were striking thirteen."
count={}
for character in message.upper():
	count.setdefault(character,0)
	count[character] = count[character] +1
pprint.pprint(count)
'''

### Tic Tac Toe##

#Data Structures

'''theBoard = {'top-L': ' ', 'top-M': ' ', 'top-R': ' ',
            'mid-L': ' ', 'mid-M': ' ', 'mid-R': ' ',
            'low-L': ' ', 'low-M': ' ', 'low-R': ' '}
def printBoard(board):
    print(board['top-L'] + '|' + board['top-M'] + '|' + board['top-R'])
    print('-+-+-')
    print(board['mid-L'] + '|' + board['mid-M'] + '|' + board['mid-R'])
    print('-+-+-')
    print(board['low-L'] + '|' + board['low-M'] + '|' + board['low-R'])
turn = "X"

for i in range(9):
    print("Turn for " + turn+ " .Move on which space?")
    move = input()
    theBoard[move] = turn
    if turn == "X":
        turn = "O"
    else:
        turn = "X"
printBoard(theBoard)'''

##Nested Dictionaries and lists

'''allGuests = {"Alice":{"apples": 5,"pretzels":12},"Bob":{"ham sandwiches": 3,"apples":2},"Carol":{"cups":3,"apple pies":1}}

def totalBrought(guests,item):
    numBrought = 0
    for k,v in guests.items():
        numBrought = numBrought + v.get(item,0)
    return numBrought

print("Number of things being brought:")
print("Apples " + str(totalBrought(allGuests,'apples')))
print("Cups " + str(totalBrought(allGuests,"cups")))
print("Cakes " + str(totalBrought(allGuests,"cakes")))
print("Ham sandwiches " + str(totalBrought(allGuests,"ham sandwiches")))
print("Apple Pies " + str(totalBrought(allGuests,"apple pies")))


while True:
    print("Enter your age?")
    age = input()
    if age.isdecimal():
        break
    print("Please enter a number for your age")

while True:
    print("Select a new password 'letters and numbers only'")
    password = input()
    if password.isalnum():
        break
    else:
        print("Passwords can only contain numbers and letters")'''

##-----String Methods----##


'''def printpicnicitems(itemDict,leftWidth,rightWidth):
    print("PICNIC ITEMS".center(leftWidth+rightWidth,"-"))
    for k,v in itemDict.items():
        print(k.ljust(leftWidth,"-") + str(v).rjust(rightWidth,"-"))

picnicitems = {"sandwiches": 4,"apples":12,"cups":4,"cookies":8000}

printpicnicitems(picnicitems,15,5)'''

## --------------------------------- Regular Expressions ------------------------------ ##

### First way of finding the phone numbers without using the regular expressions
### Trying to find and American phone number (three numbers, a hypen,three more numbers, a hypen and 4 more numbers)
### Eg 415-555-4242


'''def isPhoneNumber(text):
    if len(text) != 12:
        return False''' # not phone number sized
'''
    for i in range(0,3):
        if not text[i].isdecimal():
            return False''' # No area code
'''
    if text[3] != "-":
        return False''' # missing dash
'''
    for i in range(4,7):
        if not text[i].isdecimal():
            return False '''# not first 3 digits
'''
    if text[7] != "-":
        return False '''# missing second dash
'''
    for i in range(8,12):
        if not text[i].isdecimal():
            return False''' # missing last 4 digits
'''
    return True

print(isPhoneNumber("415-555-1234"))
print(isPhoneNumber("Hello"))



message = "Call me 415-555-1011 tomorrow, or at 415-555-9999 which is my office phone number"
foundNumber = False
for i in range(len(message)):
    chunk = message[i:i+12]
    if isPhoneNumber(chunk):
        print("Phone number found: " + chunk)
        foundNumber = True
if not foundNumber:
    print("Could not find any phone numbers.")'''


## This is with the use of regular expression notice how short the code is

'''
import re
message = "Call me 415-555-1011 tomorrow, or at 415-555-9999 which is my office phone number"
phoneNumRegex = re.compile(r"\d\d\d-\d\d\d-\d\d\d\d")
mo = phoneNumRegex.search(message)
print(phoneNumRegex.findall(message))
print(mo.group())
'''
###  Grouping with Parantheses ###

'''import re

phoneNumRegex = re.compile(r"(\d\d\d)-(\d\d\d-\d\d\d\d)")''' # Grouping the regexes Group 1 and Group 2
'''phoneNumRegex.search("My number is 415-555-4242")
mo = phoneNumRegex.search("My number is 415-555-4242")
print(mo.group(1))''' # First Group
'''print(mo.group(2))''' # Second Group

## If phone number has Parantheses around it like so (415) 555-4242

'''phoneNumRegex = re.compile(r'(\(\d\d\d\)) (\d\d\d-\d\d\d\d)')
mo = phoneNumRegex.search('My phone number is (415) 555-4242.')

print(mo.group())
areaCode,mainNumber = mo.groups()''' # multiple assignment works with the groups as well,Becouse they are tuples
'''print(areaCode)'''

### Pipe Character | 

'''batRegex = re.compile(r'Bat(man|mobile|car|bat)')
mo = batRegex.search("Batmobile has a wheel")
print(mo.group())'''

## Optional matching with the Question mark ##
# Zero or one ? , Escape character is \?

'''batRegex = re.compile(r'Bat(wo)?man')
mo = batRegex.search('The adventures of Batman') '''# Happens zero time
'''print(mo.group())
mo = batRegex.search('The Adventures of Batwoman') ''' # Happens one time and one time only
'''print(mo.group())
phoneRegex = re.compile(r'(\d\d\d-)?\d\d\d-\d\d\d\d') '''# Firts group is optional
'''mo = phoneRegex.search('My phone number is 555-1223')
print(mo.group())
'''
## Matching Zero or More with the Star 
# * symbol , Escape with \*

'''batRegex = re.compile(r'Bat(wo)*man') # Zero or More times (wo)
mo = batRegex.search('The Adventures of Batwowowowowowoman')
print(mo.group())
'''
## Matching One or More with the Plus(+)
# + symbol , Escape with \+

'''batRegex = re.compile(r'Bat(wo)+man')
mo = batRegex.search('The adventures of Batwowowowowoman')''' # Will match more
'''print(mo.group())
mo = batRegex.search('The Adventures of Batwoman') '''# Will match one
'''print(mo.group())
mo = batRegex.search('The adventures of Batman')''' # Will not Match Zero
'''
try:
    mo.group() == False
except AttributeError:
    print('it is equal to None')
'''
## Escaping examples
'''regex = re.compile(r'(\+) (\*) (\?)')
mo = regex.search('I learned about (+ * ?) regex syntax')
print(mo.group())'''

## Matching the specific Repetitions with Braces
# {x} (Exactly x), escape with \{x\}

'''haRegex = re.compile(r'(Ha){3}')''' # Exactly three times of Ha
'''mo = haRegex.search('HaHaHa I like it')
print(mo.group())
phoneRegex = re.compile(r'(\d{3}-)?(\d{3}-\d{4})')
mo = phoneRegex.search('My Phone number is 442-4444')
print(mo.group())
'''
## Greedy and Non-Greedy mathing
#By Default Python uses greedy matching ( matching the longest object)

'''greedHaRegex = re.compile(r'(Ha){3,5}')
mo=greedHaRegex.search('HaHaHaHaHa')''' # 5 times of Ha
'''print(mo.group())

nongreedyHaRegex = re.compile(r'(Ha){3,5}?')''' # Non Greedy matching with ?
'''mo=nongreedyHaRegex.search('HaHaHaHaHa') '''# Again 5 times
'''print(mo.group())

digitRegex = re.compile(r'(\d){3,5}')''' # Notice the greedy Match
'''mo = digitRegex.search('123456789') '''# it will match only 'till 5
'''print(mo.group())''' # this will be 12345

'''nongreedyDigitRegex = re.compile(r'(\d){3,5}?')''' # ? makes it non greedy
'''mo = nongreedyDigitRegex.search('123456789')''' # Will only match up to 123
'''print(mo.group())''' # this is 123

## .findall() Method

#.search() method only finds the first match and return Match Objects
#.findall() method allows us to match all matches in a string and return list of Strings
'''phoneRegex =re.compile(r'(\d{3}-)(\d{3}-\d{4})')
findAll = phoneRegex.findall('508-555-1234 and 505-445-4458 are the numbers that you may contact me')'''
# This will give a list with 2 tuples inside and inside of those tuples there are 2 items ( one area code and one main number)
'''print(findAll)'''

## Character Classes


# \d is a character class ( digit class)
# Essentially \d = r'1|2|3|4|5|6|7|8|9|0' ( its a shortcut)

## 12 days of Christmas Example

'''lyrics = "12 drummers drumming, 11 pipers piping
10 lords a leaping, 9 ladies dancing, 8 maids a milking
7 swans a swimming, 6 geese a laying, 5 gold rings
4 calling birds, 3 French hens
2 turtle doves and 1 partridge in a pear tree"

xmasRegex = re.compile(r'\d+\s\w+\s\w+')''' # more than 1 digits '\d+' a space \s and more than one word '\w+'
'''example=xmasRegex.findall(lyrics)
print(example)''' # will give a list of strings

## Making your own Character Classes
# Syntax of it is [] square brackets

'''regexObj = re.compile(r'[aeiouAEIOU]') '''# r'a|e|i|o|u' you can see latter is longer
'''vovels = regexObj.findall('Robocop eats baby food')
print(vovels) '''
# Double vowels
'''regexObj = re.compile(r'[aeiouAEIOU]{2}')''' # will only match 2 vowels next to each other
'''doubleVowels = regexObj.findall(lyrics)
print(doubleVowels)'''
# ^ symbol for negative classes
'''regexObj = re.compile(r'[^aeiouAEIOU]')''' # this will match only the letters which are outside of the list
'''constant = regexObj.findall('Robocop eats baby food')
print(constant)
'''
## Caret and Dollar Sign Characters

'''beginsWithHelloRegex = re.compile(r'^Hello')''' # has to occur at the beginnig of the string
'''beginsWithHelloRegex.search('Hello there!')''' # Will match hello
'''beginsWithHelloRegex.search('he said "Hello"') '''# will return none value

'''endsWithWorldRegex = re.compile(r'world!$')''' # has to occure at the end of the string
'''endsWithWorldRegex.search('Hello world!') '''# will match world!
'''endsWithWorldRegex.search('Hello world! and how you doin') '''# will return None

# ^both$ means pattern must match the entire string

'''allDigitsRegex = re.compile(r'^\d+$')
allDigitsRegex.search('252323523423554423')''' # will match entire number
'''allDigitsRegex.search('1324453x1235423')''' # will return None Attention at 'x'

## Wildcard Character ## or . (dot) Character
# any character except a newline

'''atRegex = re.compile(r'.at')''' # anything followed by 'at'
'''example = atRegex.findall('The cat in the hat sat on the flat mat')'''
'''print(example)''' # will return cat,sat,lat,mat ( flat didnt match becouse . character only looks for a single character)

'''atRegex = re.compile(r'.{1,2}at')''' # 1 or 2 characters and then -at
'''example = atRegex.findall('The cat in the hat sat on the flat mat')
print(example)''' # will also return 'flat' also notice the new space empty char.

## Matching everything with dot-star
#.* (wildcard and zero or more) means anything

'''string = "First Name: Erim Last Name: Serdönmez" '''# if you want to slice the First name:

'''print(string.find(':')) '''# will be 10
'''print(string.find(':') +2) '''# will be 12
'''print(string[12:])''' # will give you Erim last name:Serdönmez ( but as you can see its really long)

# Shorter version with re

'''nameRegex = re.compile(r'First Name: (.*) Last Name: (.*)')''' # look for the First Name Colon and space then whatever comes after untill Last Name colon and whatever comes after
'''print(nameRegex.findall(string))''' # will return list of tuples with Erim,Serdönmez

# dot-star will use greedy matching again you may use ? for non-greedy
# .*?

'''serve = '<To serve humans> for dinner.>'
nongreedy = re.compile(r'<(.*?)>')'''
'''print(nongreedy.findall(serve))''' # will return to serve humans
'''greedy = re.compile(r'<(.*)>')''' # greedy match
'''print(greedy.findall(serve))''' # will return to serve humans for dinner

## Matching Newlines with the dot character ##
# dot can be anything but the newline
''''
prime = 'Serve the public trust. \nProtect the innocent \nUphold the law.'

dotStar = re.compile(r'.*',re.DOTALL) '''# passing the second argument dotall to include the newline as well as everything
'''print(dotStar.findall(prime))''' # will also include \n character
#re.IgnoreCase
'''vowelRegex =re.compile(r'[aeiou]',re.IGNORECASE)''' # you dont need to type AEIOU
'''print(vowelRegex.findall('Why do you talk About the ROBOCOP so mUUUUUch?'))''' # will include the upper cases as well thanks to the re.IGNORECASE, shortcut is also re.I

## Subsitituting strings with the sub() Method ##
#They can replace it with some other texts
'''namesRegex = re.compile(r'Agent \w+')
print(namesRegex.findall('Agent Alice gave the secret documents to Agent Bob'))'''# will return the Agent's names
'''print(namesRegex.sub('REDACTED' ,'Agent Alice gave the secret documents to Agent Bob')) '''# will substitue the text REDACTED with the Agent names
# first argument is the substitue second argument is the string data
# Using \1,\2 etc in sub()

'''namesRegex = re.compile(r'Agent (\w)\w*')''' # One word character followed by zero or more other charecters (\w) is a group 1
'''print(namesRegex.findall('Agent Alice gave the secret documents to Agent Bob'))''' # Will only return A and B 
'''print(namesRegex.sub(r'Agent \1*****', 'Agent Alice gave the secret documents to Agent Bob'))

agentNamesRegex = re.compile(r'Agent (\w)\w*')
print(agentNamesRegex.sub(r'Agent \1****','Agent Alice told Agent Carol that Agent Eve knew Agent Bob was a double Agent.'))
'''
## Verbose Mode with re.VERBOSE
# re.VERBOSE will ignore the white space
'''
phoneNumRegex = re.compile(r"
(\d\d\d-)|    # So you can add comments here becouse of VERBOSE
(\(\d\d\d\))    # -or- area code with parens and no dash
(\d\d\d     # Pyhton will ignore the white space
-
\d\d\d\d)
\sx\d{2,4}  # extension, like x1234",re.VERBOSE)

phoneRegex = re.compile(r'((\d{3}|\(\d{3}\))?(\s|-|\.)?\d{3}(\s|-|\.)\d{4}(\s*(ext|x|ext.)\s*\d{2,5})?)')
'''
# Same as below notice the difference

'''phoneRegex = re.compile(r"(
    (\d{3}|\(d{3}\))?   # Area Code
    (\s|-|\.)?          #Seperator
    \d{3}               #First 3 digits
    (\s|-|\.)           #Seperator
    \d{4}               #last 4 digits
    (\s*(ext|x|ext.)\s*\d{2,5})?    # Extension
    )"",re.VERBOSE)'''


##------------------------------------Input Validation ---------------------------------##

# ask users their age and validates the input

'''while True:
    print('Enter your age')
    age = input()
    try:
        age = int(age)
    except:
        print('Please use numeric digits')
        continue
    if age < 1:
        print('Please enter a positive number')
        continue
    break

print(f'Your age is {age}.')'''

# When you run this code, you'll be prompted for your age until you enter a valid one.
# This ensures that by the time 'While' loop ends, the 'age' variable will contain a valid data type

## For un-tedious input validation we will use PyInputPlus module

'''import pyinputplus as pyip'''

# these funtcions will auto reprompt the user 'till valid data entered.

'''response = pyip.inputNum()'''  # Will ensure the input data is numerical (int)


# Just as you can pass string to input() to provide a prompt you can pass a string to a pyip function's prompt keyword argument to display a prompt.

'''response = input('Enter a number')  '''    # will accept any data
'''response = pyip.inputNum('Enter a Number')''' # will only accept a int number or float

# help(pyip.inputChoice) will display a help information about the pyip.inputChoice function
# for more information pyinputplus.readthedocs.io

## The min,max,greaterThan and lessThan keyword Arguments.

# inputNum(), inputInt() and inputFloat() will accept int and float numbers
# we also have min,max,greaterThan and lessThan keyword arguments for specifying a range of valid values.

'''response = pyip.inputNum('Enter a number: ',min=4) '''# will accept minumum '4' as a number
'''response = pyip.inputNum('Enter a number: ',greaterThan=4)''' # will accept greater than '4' but not 4
'''response = pyip.inputNum('>',min=4,lessThan=6)''' # will accept min '4' and less than '6' but not '6'

## The Blank Keyword Argument ##
# By Default, blank input is not allowed unless blank keyword argument is set to True:

'''response = pyip.inputNum('Enter num:')''' # will not accept a blank value
'''response = pyip.inputNum('Enter a num:',blank=True)''' # will accept a blank value

## The limit,timeout and default Keyword Arguments ##

'''response = pyip.inputNum('Enter a number: ',limit=2)''' # will ask twice before giving up
'''response = pyip.inputNum('Enter a  Number: ',timeout=10) '''# will wait 10 secs before giving up
'''response = pyip.inputNum('Enter a number: ',limit=2,default='You only have 2 chances')
print(response)''' # default value will be stored in the 'response' variable

## The allowRegexes and blockRegexes keyword Arguments ##

'''response = pyip.inputNum(allowRegexes=[r'(I|V|X|L|C|D|M)+',r'zero'])''' # will allow roman characters as a number 'XLII' will work even though its not a standart int number
'''response = pyip.inputNum(blockRegexes=[r'[02468]$'])''' # will not accept even numbers '$'means at the end of the input data
'''response = pyip.inputStr(allowRegexes=[r'caterpillar','category'],blockRegexes=[r'cat'])''' # will not accept cat or catwoman but will accept caterpillar and category
# More info at pyinputplus.readthedocs.io

## Passing a Custom validation Function to inputCustom() ##
# ask for user input to enter a series of digits that adds up to 10
#   Accepts a single string argument of what the user entered
#   Raises and exception if the string fails validation
#   Return None(or has no return statement) if inputCustom() should return the string unchanged
#   Returns a non-None value if inputCustom() should return a different string from the one user entered
#   Is passed as the first argument to inputCustom()

## Creating our own addUpToTen() Function

'''def addsUpToTen(numbers):
    numbersList = list(numbers)
    for i, digit in enumerate(numbersList):
        numbersList[i] = int(digit)
    if sum(numbersList) != 10:
        raise Exception('The digits must add up to 10, not %s'%(sum(numbersList)))
    return int(numbers)''' # return an int form of numbers.

'''response = pyip.inputCustom(addsUpToTen)'''

##-------------------------------- FILES ----------------------------------------------##

# To extract each attribute from the file path (Anchor,Parent,Name)
'''from pathlib import Path
import os
'''

'''p = Path('C:/Users/devil/spam.txt')''' # Notice the forward slash instead of backslash
'''print(p.anchor)'''     # will return C:\
'''print(p.parent)'''     # will return C:\Users\devil # This is a Path Object not a string
'''print(p.name)  '''     # will return spam.txt
'''print(p.suffix) '''    # will return .txt
'''print(p.stem)'''       # will return spam
'''print(p.drive)  '''    # will return 'C:'

'''print(Path.cwd())'''

# Finding the file sizes

'''totalSize = 0
for filename in os.listdir('C:\\Windows\\System32'):
    totalSize = totalSize + os.path.getsize(os.path.join('C:\\Windows\\System32',filename))'''
    
'''As I loop over each filename in the C:\Windows\System32 folder, the totalSize variable is incremented by the size of each file. Notice how when I call os.path.getsize(), I use os.path.join() to join the folder name with the current filename. The integer that os.path.getsize() returns is added to the value of totalSize. After looping through all the files, I print totalSize to see the total size of the C:\Windows\System32 folder.'''   



# Using the glob() method

'''p = Path('C:/Users/devil/desktop')
print(p.glob('*'))  '''# will return generator object

'''print(list(p.glob('*')))'''   # will make a list from the generator

# (*) stands for 'multiple of any characters' so p.glob('*') returns a generator of all files in the path stored in p.

# like with regexes you can create complex expressions:

'''print(list(p.glob('*.txt')))  '''      # will list all txt files

# In contrast to (*), '(?)' stands for any single character

'''print(list(p.glob('project?.docx'))) '''   # will return 'project1.docx' or 'project5.docx' but will not return 'project10.docx' 'couse ? stands for single character

# Combining the asterix and question mark

'''print(list(p.glob('*.?x?')))'''    # will return files with any name and any three character extension (middle charecter is x)


# By picking out files with specific attributes, the glob() method lets you easily specify the files in a directory you want to perform some operation on. You can use a for loop to iterate over the generator that glob() returns:


'''for textFilePathObj in p.glob('*.txt'):
    print(textFilePathObj)'''  # prints the Path object as a string
    # Do something with the text file

# If you want to perform some operation on every file in a directory, you can use either os.listdir(p) or p.glob('*').

## THE FILE READING/WRITING PROCESS ##
# The pathlib module has read_text() and write_text() methods

'''from pathlib import Path
from typing import SupportsBytes'''

'''p = Path('spam.txt')'''
'''print(p.write_text('Hello World')) ''' # will create and write this string into a file
'''print(p.read_text())    '''            # will read the file
                                    # file location is at cwd

## Opening Files with open() Function

'''helloFile = open(Path.home()/'hello.txt')
helloFile = open('C:\\Users\\devil\\hello.txt')
'''
## Reading the contents of Files 

'''content = helloFile.read() '''# will read the file and store it in a variable
'''print(content)'''

'''sonnetFile = open(Path.home()/'sonnet29.txt')'''
'''print(sonnetFile.readlines())   '''# will return a list value of each line

## Writing to Files

'''baconFile = open('bacon.txt','w')''' # will create a new file called 'bacon' and will open it in 'write mode'
'''baconFile.write('Hello World!\n')''' # will write the string to the file
'''baconFile.close()    '''             # Close the file
'''baconFile = open('bacon.txt','a') '''# Open the file in append mode
'''baconFile.write('Bacon is not a vegetable.')''' # will write the string to the end of the file
'''baconFile.close()  '''               # close the file
'''baconFile = open('bacon.txt')  '''   # Open the file in read-only mode
'''content = baconFile.read() '''       # store the file contents in a variable
'''baconFile.close()     '''            # close the file
'''print(content) '''                   # display the contents from the file

## Saving variables with the Shelve Module ##

'''import shelve'''

'''shelFile = shelve.open('mydata') '''   # call shelve.open and pass it a filename
'''cats = ['Zophie','Pooka','Simon']'''
'''shelFile['cats'] = cats  '''           # Store the returned shelf value in a variable
'''shelFile.close()
'''
'''shelFile = shelve.open('mydata')'''
'''print(type(shelFile))  '''             # class 'shelve.DbfilenameShelf'
'''print(shelFile['cats'])  '''           # [Zophie,Pooka,Simon]
'''shelFile.close()'''

'''shelFile = shelve.open('mydata')'''
'''print(list(shelFile.keys())) '''       # will display ['cats']
'''print(list(shelFile.values()))  '''    # will display ['Zophie','Pooka','Simon']
'''shelFile.close()
'''
## Saving variables with the pprint.pformat() Function ## 
#  Say you have a dictionary stored in a variable and you want to save this variable and its contents for future use. Using pprint.pformat() will give you a string that you can write to a .py file. This file will be your very own module that you can import whenever you want to use the variable stored in it.

'''import pprint
cats = [{'name':'Zophie','desc':'chubby'},{'name' : 'Pooka','desc':'fluffy'}]
print(pprint.pformat(cats))

fileObj = open('myCats.py','w')
fileObj.write('cats='+pprint.pformat(cats)+'\n')''' # we will keep this dict even after we close the shell
'''fileObj.close()'''

## Organizing Files ##

# Shutil module and Copying

'''import shutil,os
from pathlib import Path

p = Path.home()
shutil.copy(p/'hello.txt',p/'some_folder')'''  # copies file to the folder.The return value is the path of the newly copied file.
'''shutil.copy(p/'hello.txt',p/'some_folder/hello2.txt') '''# also copies the file but gives it a new name.

'''shutil.copytree(p/'some_folder',p/'some_folder_backup')''' # will copy the entire folder and every folder and file inside to a different location

# Shutil module and Renaming and moving files and folders

'''import shutil

shutil.move('C:\\users\\devil\\some_folder','C:\\users\\devil\\desktop')''' # will move that file or folder to a new location
'''shutil.move('C:\\users\\devil\\hello.txt','C:\\users\\devil\\desktop\\some_folder\\hello3.txt')''' # will move that file to a location and will rename it.

## Walking a directory tree ## 
'''import os,shutil


for folderName,subFolders,fileNames in os.walk('c:\\users\\devil\\delicious'):
    print('The current folder is ' + folderName)

    for subFolder in subFolders:
        print('Subfolder of ' + folderName + ' : '+ subFolder)

    for filename in fileNames:
        print('File inside ' + folderName + ': '+ filename)

    print()
'''

## Compressing files with the zipfile module
# Reading a Zipfile

'''import zipfile,os
from pathlib import Path

p = Path.home()    ''' # C:\users\devil

'''exampleZip = zipfile.ZipFile(p/'example.zip')
print(exampleZip.namelist()) '''   # will return the contents of the zipfile as a list of strings

'''spamInfo = exampleZip.getinfo('spam.txt')
print(spamInfo.file_size)'''       # will return the actual size

'''print(spamInfo.compress_size) '''  # will return the compressed size

'''print(f'Compressed file is {round(spamInfo.file_size/spamInfo.compress_size,2)}x smaller !!')'''
# will return ' Compressed file is 3.63x smaller!!'
'''exampleZip.close()'''

# Extracting from zip files

'''import zipfile,os
from pathlib import Path

p = Path.home()

exampleZip = zipfile.ZipFile(p/'example.zip')'''
'''exampleZip.extractall() '''# will extract at cwd, extractall('C:\\users\devil\\delicious') will extract to that particular folder


'''exampleZip.extract('spam.txt')''' # will extract to cwd
'''exampleZip.extract('spam.txt','C:\\users\\devil\\delicious')''' # will extract to that folder, if not exist python will create it 

'''exampleZip.close()'''

## Creating and Adding to ZipFiles.

'''import zipfile

newZip = zipfile.ZipFile('new.zip','w')
newZip.write('C:\\users\\devil\\spam.txt',compress_type=zipfile.ZIP_DEFLATED)
newZip.close()'''

## ---------------------------------- DEBUGGING -----------------------------------------##
"""
*********************
*                   *
*                   *
*                   *
*********************

"""
'''def boxPrint(symbol,width,height):
    print(symbol * width)

    for i in range(height-2):
        print(symbol + (' ' * (width-2))+ symbol)

    print(symbol * width)'''


#boxPrint('-',10,5) # this will print the box as above no problems whatsoever
#boxPrint('--',10,5) # However this will be bugged 'couse of double string entry

'''def boxPrint_debugged(symbol,width,height):
    if len(symbol) != 1:
        raise Exception('Symbol must be a single character string.')
    if width <= 2: ## careful with the len(int)
        raise Exception('Width must be greater than 2.')
    if height <= 2:
        raise Exception('Height must be greater than 2.')

    print(symbol* width)

    for i in range(height-2):
        print(symbol + (' '*(width-2)) + symbol)

    print(symbol * width)'''


#boxPrint_debugged('--',10,5) # will raise the exception number 1
#boxPrint_debugged('-',1,1)   # will raise the second exception
'''for sym,w,h in (('*', 4, 4), ('O', 20 , 5), ('-', 1, 3), ('ZZ', 3, 3)):
    try:
        boxPrint_debugged(sym,w,h)
    except Exception as err:
        print(' An Exception happened: ' + str(err))'''

# The traceback.format_exc() Function to get traceback call error message as a string

'''import traceback

try:
    raise Exception('this is the error message...')
except:
    errorFile = open('error_log.txt','a')
    errorFile.write(traceback.format_exc())
    errorFile.close()
    print('The traceback info was written in error_log.txt')'''

# Above example will create an append mode .txt file in cwd and will store the traceback log in it.

# Assertions and the assert Statment

'''market_2nd = {'ns': 'green','ew':'red'}

def switchLights(intersection):
    for key in intersection.keys():
        if intersection[key] == 'green':
            intersection[key] = 'yellow'
        elif intersection[key] == 'yellow':
            intersection[key] = 'red'
        elif intersection[key] == 'red':
            intersection[key] = 'green'
    assert 'red' in intersection.values(), 'neither light is red!' + str(intersection)'''  # Makes sure that there is at leas one red light, so the cars will not crash into each other

'''print(switchLights(market_2nd))''' # will switch red to green and green to yellow so outcome will be green and yellow therefore no red light which will raise the Assertion error.


'''ages = [26,57,92,54,27,22,15,80,47,73]
ages.reverse()  
ages.sort()
print(ages)
assert ages[0] <= ages[-1],"Assert that the first age is <= the last age"''' 
# assert will be returned 'couse the list is sorted reverse, if sort() correctly which means True Condition has been fulfilled the assert code will skip and will do nothing.

## LOGGING ## 

'''import logging

logging.basicConfig(filename='myProgram.txt',level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')'''
# This is the set up code for logging in Python

# Example: Buggy Factorial Program

'''def factorial(n):
    total = 1
    for i in range(n+1):
        total *= i
    return total
'''
'''print(factorial(10))''' # will return Zero (0) which means theres a bug in somewhere inside the function code

# logging.debug() Function

'''logging.disable(logging.CRITICAL)''' # will get rid of the logging.debug messages as seen below.
'''logging.debug('Start of program')


def factorial_debugged(n):
    logging.debug('Start of factorial (%s)' % (n))
    total = 1
    for i in range(n+1):
        if i == 0:
            continue
        total *= i
        logging.debug(' i is %s, total is %s '%(i,total))

    logging.debug('Return value is %s' %(total))    
    return total

print(factorial_debugged(4))

logging.debug('End of program')'''

# logging.disable(logging.CRITICAL) will get rid of the logging.debug messages

# Logging to a Text file logging.basicConfig(FILENAME='MYPROGRAM.TXT,level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s') check out the first entry on that line
# filename='myprogram.txt, will create a text file at cwd which can log your messages inside

# Logging Levels

'''import logging
logging.basicConfig(level=logging.DEBUG,format='%(asctime)s- %(levelname)s- %(message)s')'''
# lowest level
'''logging.debug('some debugging details. ')

logging.info('The logging module is working. ')

logging.warning(' An error message is about to be logged. ')

logging.error('  An error has occured. ')

logging.critical('The program is unable to recover. ')'''
# highestlevel

# Using debugger

'''print('Enter the first number to add: ')
first = input()
print('Enter the second number to add: ')
second = input()
print('Enter the third number to add: ')
third = input()
print('the sum is ' + first + second + third)'''

'''import random

heads = 0

for i in range(1,1001):
    if random.randint(0,1) == 1:
        heads += 1
    if i ==500:
        print('halfway done.')

print('heads come up ' + str(heads) + ' times')'''

## ------------------------- Web Scraping ----------------------------------------------##


'''import webbrowser,sys,pyperclip
'''
# sys.argv System variables will use sys.argv function ('mapit.py','870','Valencia','St.') 


# Check if command line arguments were passed

## If you run the program by entering this into the command line...
# mappit 870 Valencia St, San Francisco, CA 94110
# sys.argv variable will contain this LIST value;
# ['mappit','870','Valencia','St.','San','Francisco',',','CA','94110']

'''if len(sys.argv) > 1:'''
    # ['mappit.py,'870','Valencia','st.'] -> '870 Valencia st.'
    #address = ' '.join(sys.argv[1:]) # will only contain ('870' and after)
#else:
#    address = pyperclip.paste() # Whatever text is n the clipboard and returns a string which we store to address variable

# https://www.google.com/maps/place/870+Valencia+St,+San+Francisco,+CA+94110,+Amerika+Birle%C5%9Fik+Devletleri/@37.7589621,-122.4238514,17z/data=!3m2!4b1!5s0x808f7e3dae6df76b:0x6240f16ee9572080!4m5!3m4!1s0x808f7e3dae0fc797:0x26acf7c8a5797e94!8m2!3d37.7589579!4d-122.4216627
# as you can see this url for the address is too long
# https://www.google.com/maps/place/<ADDRESS>
# Because google is smart enough to understand we can shorten the url as above. After that we can just use the 'address' variable as an extension for that url as below;

#webbrowser.open('https://www.google.com/maps/place/' + address)

## To continue the above program faster I will create a batch file in order to doing so I will need to create a new file in 'C:\\Users\\devil\\myPythonScripts' so just follow from there.
## Then I will just use windows + R key to run mappit <sys.argv()> 

## Downloading from the Web with the Requests module. ##

'''from os import error
import requests

res = requests.get('https://automatetheboringstuff.com/files/rj.txt')'''

# requests.get() will take the string URL and will return a 'response Object'

'''print(res.status_code) '''# if none value returns (or nothing happens) everything is ok.

# you may also try:
'''if res.status_code == requests.codes.ok:# this will return a Boolean value. Because requests.codes.ok is a variable which contains the code '200'.
    print('OK.')'''

## For more info on HTTP status codes ;
# en.wikipedia.org/wiki/List_of_HTTP_status_codes.

'''print(len(res.text)) '''# will show you the ' length of characters' in that file.
# Checking for Errors: #
'''print(res.raise_for_status())''' # will check if the expression was succesfull or not as below
'''try:
    badRes = requests.get('http://automatetheboringstuff.com/random.txt')''' # This doesn't exists so it will raise an exception.
'''except error:
    res = requests.get('https://automatetheboringstuff.com/files/rj.txt')''' # This is correct so it will do this instead.

# Saving Downloaded files to the hard Drive #

# you need to open the file in Binary write mode as below to maintain the Unicode encoding.
'''playFile = open('RomeoAndjJuliet.txt', 'wb')'''
'''try:
    playFile.write(res)''' # will not work because its a response object and not a string of text.
'''except TypeError:
    # iter_content() method will return 'bytes' data type in chuncks which we can use it to write in a file.
    
    for chunk in res.iter_content(100000):
        playFile.write(chunk)

playFile.close()'''

# For more info on Unicode Encodings visit;
#   joelonsoftware.com/articles/Unicode.html
#   nedbatchelder.com/text/unipain.html

##  Here is the complete process for downloading and saving a file:
    # 1. Call requests.get() to download the file.
    # 2. Call open() with 'wb' to create a new file in write binary mode.
    # 3. Loop over the 'Response' object's iter_content() method.
    # 4. Call write() on each iteration to write the content to the file.
    # 5. Call close() to close the file.

# For more info on requests module visit requests.readthedocs.org

## HTML ##

# Resources for Learning HTML

#   https://developer.mozilla.org/en-US/learn/html/
#   https://html.dog.com//guides/html/beginner/
#   https://www.codeacademy.com/learn/learn-html

# Beatiful Soup module
#import bs4,requests
'''
def getAmazonPrice(productUrl):
    res = requests.get(productUrl)
    res.raise_for_status()

    soup = bs4.BeautifulSoup(res.text,'lxml')
    elems=soup.select('#s0-0-18-5-11-26-97 > div > span > div')
    return elems[0].text.strip()



price = getAmazonPrice('https://www.ebay.com/p/28023709776?iid=265235066301')    

print(f'the price is {price}' )
'''

'''res = requests.get('https://forecast.weather.gov/MapClick.php?lat=37.789690000000064&lon=-122.39571499999994#.YPXDqugzaUk')
res.raise_for_status()

soup = bs4.BeautifulSoup(res.text,'lxml')
elems = soup.select('#detailed-forecast-body > div:nth-child(1) > div.col-sm-10.forecast-text')
print(f'The weather in San francisco today is {elems[0].text} ')'''

'''res = requests.get('https://nostarch.com')  # will download the main page

res.raise_for_status()

noStarchSoup = bs4.BeautifulSoup(res.text, 'html.parser')   # passes the text attribute of the response to bs4.BeautifulSoup() 

print(type(noStarchSoup))   # BeautifulSoup Object stored in variable 'noStarchSoup'

exampleFile = open('exampleHTML.html')   # this is a file from my hard disk
exampleSoup = bs4.BeautifulSoup(exampleFile,'lxml') # same way it can be open with bs4
print(type(exampleSoup))    # again it will be BeautifulSoup object

elems = exampleSoup.select('#author')   # elems is a list of tag objects
print(type(elems)) , print(len(elems)), print(type(elems[0]))'''
# theres one TAG object in the list, there was one match

'''print(str(elems[0])) '''   # the Tag object as a string.
# Passing the element to str() returns a string with starting and closing tags and the element's text
'''print(elems[0].getText()) , print(elems[0].attrs)'''
# calling .getText() on elements returns the element's text or inner HTML
# Attrs gives us a dictionary with the element's attribute, 'id' and the value of the id attribute, 'author'


# you can also put <p> elements from the BeautifulSoup object
'''pElems = exampleSoup.select('p')
print(len(pElems))      # 3 matches
print(pElems[0]), print(pElems[0].getText())
print(str(pElems[1])) , print(pElems[1].getText())
print(str(pElems[2])), print(pElems[2].getText())'''

# using str() will show each element as a string, getText() will show you the inner HTML

# using get() method to get data from element's attributes

'''soup = bs4.BeautifulSoup(open('exampleHTML.html'),'lxml')
spanElem = soup.select('span')[0]

print(str(spanElem)) , print(spanElem.get('id'))
print(spanElem.get('some_nonexistent_addr') == None) , print(spanElem.attrs)'''

# Here we use select() to find any <span> elements and then store the first matched elenment in spanElem. Passing the attribute name 'id' to get() returns the attribute's value 'author'
## For more info about BeautifulSoup go to: https://www.crummy.com/software/BeautifulSoup/bs4/doc/

## Controlling the Browser with the Selenium Module ##

# For mnore info on user-agent strings visit : 'https://www.whatsmyua.info/'

# Starting a Selenium controlled Browser

'''from selenium import webdriver'''

#browser = webdriver.Firefox()
#browser = webdriver.Chrome()

'''print(type(browser))'''    # Webdriver data type

'''browser.get('https://inventwithpython.com')''' # will direct browser to the URL

## Finding Elements on the Page ##

'''try:
    elem = browser.find_element_by_class_name('cover-thumb')
    print('Found <%s> element with that class name!'%(elem.tag_name))
except:
    print('Was not able to find an element with that name.')'''

## Clicking the Page ##
# with click() method , will get the webelement object for the <a> element with the text 'Read it Online for Free', and then simulates clicking that <a> element.

'''linkElem = browser.find_element_by_link_text('Read Online for Free')
print(type(linkElem))
linkElem.click()  '''  # Follows the "Read Online for Free" link

## Filling out and Submitting Forms ##
# will go to the website 'login.metafiler.com' and will fill the text areas for user name and passwords. Finally will click on the submit button.
# Calling submit() method on any element will have the same result as clicking the Submit button for the form that element is in.(You could have just as easily called emailElem.submit(), and the code would have done the same thing.)

'''browser = webdriver.Firefox()

browser.get('https://login.metafilter.com')

userElem = browser.find_element_by_id('user_name')
userElem.send_keys('SketchQ')    # Your user name here

passwordElem = browser.find_element_by_id('user_pass')
passwordElem.send_keys('123456')
passwordElem.submit()'''

## Sending Special Keys ##
# selenium.webdriver.common.keys module but shorter version is ;

'''from selenium.webdriver.common.keys import Keys'''
# the <html> tag is the base tag in HTML files: the full content of the HTML file is enclosed withing the <html> and <\html> tags.
'''
browser = webdriver.Firefox()
browser.get('https://nostarch.com')
htmlElem = browser.find_element_by_tag_name('html')
htmlElem.send_keys(Keys.END)    # Scrolls to bottom
htmlElem.send_keys(Keys.HOME)   # Scrolls to top'''

## More information on Selenium ##
# visit 'https://selenium-python.readthedocs.org/'

## ---------------------------- Working with Excel SpreadSheets --------------------------##

'''import openpyxl
import os

from openpyxl.utils.cell import column_index_from_string
'''
# Opening an Excel file with openpyxl   
'''workbook = openpyxl.load_workbook('example.xlsx')'''
# Use openpyxl.load_workbook() to open the excel file.
'''print(type(workbook))'''
# it takes a file name and returns a Workbook object just like File object represents an opened text file.

# Getting Sheets from the Workbook

'''print(workbook.sheetnames)'''  # Theworkbook's sheets names.
'''sheet =workbook['Sheet1']  ''' # Get a sheet from the workbook.
'''print(type(sheet))         ''' # Worksheet object
'''print(sheet.title)          '''# Get the sheet's title as a string
'''anotherSheet = workbook.active ''' # Get the active sheet
'''print(anotherSheet)'''

# Getting Cells from the Sheets.

'''cell = sheet['A1']  '''# Get a cell from the sheet.
'''print(cell.value)   '''# Get the value from the cell
'''c = sheet['B1']     '''# get another cell from the sheet
'''print(c.value)
print('Row %s,Column %s is %s'%(c.row,c.column,c.value))'''    # get the row,column, and value from the cell    
'''print('Cell %s is %s'%(c.coordinate,c.value))   '''# get coordinate and value of the cell

'''print(sheet.cell(row=1,column=2))   '''# justlike sheet[B1] you may also use (row=1,column=2)
'''print(sheet.cell(row=1,column=2).value)
print(sheet['C1'].value)
'''
'''for i in range(1,8):    # Go through every other row:
    print(i,sheet.cell(row=i,column=2).value)

for i  in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)
'''
'''print(sheet.max_row) '''   # print the max row
'''print(sheet.max_column)''' # print max column (will return a integer value)

# Converting between Column Letters and Numbers
'''from openpyxl import utils,cell'''

'''print(openpyxl.utils.cell.get_column_letter(2))'''
'''
openpyxl.utils.cell.column_index_from_string()        # To convert from letters to numbers
openpyxl.utils.cell.get_column_letter()               # To convert numbers to letters.
'''
'''
print(openpyxl.utils.cell.get_column_letter(1))       # Translate column 1 to a letter.
print(openpyxl.utils.cell.get_column_letter(2))
print(openpyxl.utils.cell.get_column_letter(1000))
print(openpyxl.utils.cell.get_column_letter(sheet.max_column))

print(column_index_from_string('A'))                  # Get A's number.
print(column_index_from_string('AA'))               
'''
# Getting Rows and Columns From the Sheets

'''print(sheet['A1:C3'])'''   # Get all cells from A1 to C3. 

# Above we specify we want 'Cell' objects in the rectangular area from A1 to C3, and we get a 'Generator' object ontaining the 'Cell' objects in that area.
# We return a cell objects in a Tuple to visiualize the generator object
'''
for rowOfCellObjects in sheet['A1:C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate,cellObj.value)
    print('---END of ROW---')
'''
# To print the values of each cell in the area we use two for loops.
# the outer loop goes over each row in the slice
# the nested inner for loop goes through each cell in that row.

# To access the values of cells in a particular row or column, you can also use a Worksheet object's rows and columns attribute. These attributes must be converted to lists with list() function before we can us square brackets and an index with them.
'''
print(list(sheet.columns)[1])   # Get second column's cells.
print(list(sheet.rows)[1])
for listObj in list(sheet.columns)[1]:
    print(listObj.value)
    '''
# For more info about openpyxl visit ==> https://openpyxl.readthedocs.org/

## Writing Excel Documents ## 
'''
import openpyxl

wb = openpyxl.Workbook()        # Call this function to create a new blank Workbook Object
print(wb.sheetnames)            # It starts with one sheet
sheet = wb.active
sheet.title                     # name of the sheet or the title of the sheet
sheet.title = 'Spam Bacon Eggs Sheet'   # will change the name of the sheet
print(wb.sheetnames)
wb.save('example_copy.xlsx')      # Will save the workbook to your harddrive
wb.create_sheet()               # Will add a new sheet (default name Sheet 1)
wb.create_sheet(index=0,title='First Sheet')    # will create a new sheet called 'First Sheet' at index point 0 which is the begining.
wb.create_sheet(index=2,title='Middle Sheet')
print(wb.sheetnames)
wb.save('example_copy.xlsx')
del wb['Sheet']                 # will delete the names sheet
del wb['Spam Bacon Eggs Sheet']
print(wb.sheetnames)
wb.create_sheet(index=2,title='Final Sheet')
wb.save('example_copy.xlsx')
# Writing Values to Cells.

sheet = wb.active
sheet['A1'] = 'Hello There'
sheet['A2'] = 'How you doing??'
sheet['B1'] = ' I\'m fine'
sheet['B2'] = ' Hellloooo'
wb.save('example_copy.xlsx')'''

## Setting the Font Style of Cells 

'''import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24font = Font(size=24,italic=True)    # Create a font.
sheet['A1'].font = italic24font             # Apply the font to A1
sheet['A1'] = 'Hello, World!'
wb.save('styles.xlsx')

fontObj = Font(name='Times New Roman',bold=True)
sheet['B1'].font = fontObj
sheet['B1'] = 'Times new Roman Bold'

fontObj2 = Font(size=24,italic=True)
sheet['C1'].font = fontObj2
sheet['C1'] = '24 pt Italic'

wb.save('styles.xlsx')'''

## Formulas 
'''
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 500
sheet['A3'] = 255
sheet['A4'] = '=SUM(A1:A3)'
sheet['A4'].font = Font(size=24,bold=True,italic=True)
wb.save('writeFormula.xlsx')'''

## Adjusting Rows and Columns

'''import openpyxl
wb = openpyxl.Workbook()
sheet= wb.active
sheet['A1'] = 'Tall Row'
sheet['B2'] = 'Wide Column'
# set the height and width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')'''

## Merging and Unmerging Cells

'''import openpyxl
from openpyxl.styles import Font

wb =openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')  # Merge all these cells
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')  # Merge these two cells.
sheet['C5'] = 'Two merged Cells.'
sheet['A1'].font = Font(bold=True,size=24)
sheet['C5'].font = Font(italic=True)
wb.save('merged.xlsx')

sheet.unmerge_cells('A1:D3')    # Unmerge these cells
sheet.unmerge_cells('C5:D5')    
wb.save('unmerged.xlsx')'''

## Freezing Panes

"""import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'   # Freeze the rows above A2
wb.save('freezeExample.xlsx')"""

## Charts 


'''import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1,11):   # Create some data in column A
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet,min_col=1,min_row=1,max_col=1,max_row=10)
seriesObj = openpyxl.chart.Series(refObj,title='First Series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj,'C5')
wb.save('sampleChart.xlsx')
'''

## ----------------------- Chapter 14 Working with Google Spreadsheets -------------------##

# For more info on EZSheets follow:
    # https://ezsheets.readthedocs.io

'''from re import S
import ezsheets'''

# Getting Spreadsheet object from google.spreadsheet using the ID of the spreadsheet

'''ss = ezsheets.Spreadsheet('1P7TTVPgjiVTj9OfjVflSUtQsXYNAj0C_ZC8LWCvpzTY')

print(ss)

print(type(ss))'''

# Getting ss object using the Full URL

'''ss = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/1P7TTVPgjiVTj9OfjVflSUtQsXYNAj0C_ZC8LWCvpzTY/edit#gid=0')
print(ss)
print(type(ss))'''

# Getting ss object using the Tittle of the Spreadsheet

'''ss = ezsheets.Spreadsheet('Quickstart')
print(ss)
print(type(ss))'''

# Creating a new blank Spreadsheet using the ezsheets.createSpreadsheet() Function

'''ss = ezsheets.createSpreadsheet('Tittle of my New Spreadsheet')
print(ss)
print(ss.title)'''

# Uploading your Spreadsheet with ezsheets.upload() Function

'''ss = ezsheets.upload('multiplication_table_999.xlsx')
print(ss)
print(ss.title)'''

# Getting your spreadsheet from google sheets as a dictionary as keys are ID and values are Titles using the ezsheets.listSpreadsheets() Function

'''ss = ezsheets.listSpreadsheets()
print(ss)
print(type(ss))'''

## Spreadsheet Attributes

# .title Attribute for the spreadsheet's title.

'''ss = ezsheets.Spreadsheet('1zIM77SkYyFnlriAiB1QPSBIFbAIt8us5cD8yT6a9S3Q')

print(ss.title)
'''
# Changing the title attribute

'''ss.title = 'Class Data'

print(ss.title)
'''
# Getting the ID attribute from the spreadsheet(This is a read-only attribute)

'''print(ss.spreadsheetId)'''

# Getting the URL attribute of the ss

'''print(ss.url)'''

# Getting all the names of the sheets objects of that ss

'''print(ss.sheetTitles)'''

# The sheet objects in that ss (in order with row and column counts as well as sheetID's)

'''print(ss.sheets)'''

# The first Sheet object in this Spreadsheet.

'''print(ss[0])'''

# You can also get the sheets by their titles

'''print(ss['Sayfa1'])
'''
# You can delete the sheets with del statement as below
'''
try:
    del ss[0]
    print(ss.sheetTitles)
except ValueError:
    print('No sheet has been found\n You cannot delete the sheet with only one sheet')'''

# If someone changes the spreadsheet you can update it with .refresh
# This will also update the data inside the spreadsheet object

'''print(ss.refresh)'''

## Downloading and Uploading Spreadsheets


'''ss = ezsheets.Spreadsheet('1zIM77SkYyFnlriAiB1QPSBIFbAIt8us5cD8yT6a9S3Q')'''

# Downloads the spreadsheet as an Excel file.

'''print(ss.downloadAsExcel('title_here.xlsx'))
'''
# Downloads the spreadsheet as an OpenOffice file.

'''print(ss.downloadAsODS('title_here.ods'))'''

# Download as CSV file

'''print(ss.downloadAsCSV('title_here.csv'))'''

# Download as a TSV file

'''print(ss.downloadAsTSV('title_here.tsv'))'''

# Download as PDF file

'''print(ss.downloadAsPDF('title_here.pdf'))'''

# Download as a ZIP of HTML files

'''print(ss.downloadAsHTML('title_here.html'))'''

# Note that CSV and TSV formats can only contain one sheet, therefore if you need to download other sheets you'll need change sheet objects index attribute to 0 

## Deleting Spreadsheets

# Create the spreadsheet
'''ss = ezsheets.createSpreadsheet('delete me')'''
# Confirm that we've created the spreadsheet
'''print(ss)'''
# Delete the spreadsheet
'''ss.delete()'''
# Confirm the delete was ok
'''print(ezsheets.listSpreadsheets())'''
# Permanent deletion
'''ss.delete(permanent=True)'''

## Sheet Objects

'''ss = ezsheets.Spreadsheet('1zIM77SkYyFnlriAiB1QPSBIFbAIt8us5cD8yT6a9S3Q')'''

# The Sheets object in this Spreadsheet, in order.
'''print(ss.sheets)'''
# Note that this is a tuple object
'''print(type(ss.sheets))'''
# Get the first Sheet object in this Spreadsheet
'''print(ss.sheets[0])'''
# Note that this is a Sheet object
'''print(type(ss.sheets[0]))'''
# Also gets the first sheet object in this Spreadsheet
'''print(ss[0])
print(type(ss[0]))'''
# The titles of all the Sheet objects in this Spreadsheet.
'''print(ss.sheetTitles)'''
# Sheets can also be accessed by title
'''print(ss['Sayfa1'])
print(type(ss['Sayfa1']))'''

## Reading and Writing Data

# Creating a new blank Spreadsheet
'''ss = ezsheets.createSpreadsheet('My Spreadsheet')'''
# Getting the first sheet in this Spreadsheet
'''sheet = ss[0]'''
# Checking the title of that first sheet
'''print(sheet.title)'''
# Setting the values in cells
'''sheet['A1'] = 'Name'
sheet['B1'] = 'Age'
sheet['C1'] = "Favorite movie"'''
# Reading the values in those cells
'''print(sheet['A1'])
print(sheet['B1'])
print(sheet['C1'])'''
# Empty cells returns a blank string
'''print(sheet['A2'])
print(type(sheet['A2']))'''
# Column 2 and Row 1 is the same as B1.
'''print(sheet[2,1])'''
# Setting values to another cells
'''sheet['A2'] = 'Erim'
sheet['B2'] = '30'
sheet['C2'] = "Lord of the Rings"'''

## Column and Row Adressing

# Converts string addres in to a tuple adress
'''print(ezsheets.convertAddress('A2'))'''
# or Vice versa
'''print(ezsheets.convertAddress(1,2))'''
# You can get the letter for columns
'''print(ezsheets.getColumnLetterOf(2))'''
# or Vice versa
'''print(ezsheets.getColumnNumberOf('C'))'''
# More examples
'''print(ezsheets.getColumnNumberOf('ZZD'))
print(ezsheets.getColumnLetterOf(1555))'''

## Reading and Writing Entire Columns and Rows

#ss = ezsheets.upload('produceSales.xlsx')
'''ss = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/19S_1jDXh_aLs6tSGvzhR1p6Ycc4uJZnOjgezMQ5C4Fs/edit#gid=955048666')'''
# We also need a sheet object to use the sheet methods
'''sheet = ss[0]'''
# The first row is 1 not 0
'''print(sheet.getRow(1))
print(sheet.getRow(2))

columnOne = sheet.getColumn(1)'''
# Both are the same
'''print(sheet.getColumn(1))
print(sheet.getColumn('A'))'''
# Writing an entire row
'''print(sheet.getRow(3))
sheet.updateRow(3,['Pumpkin','11.50','20','230'])
print(sheet.getRow(3)) '''
# Writing an entire column
'''columnOne = sheet.getColumn(1)
for i,value in enumerate(columnOne):
    columnOne[i] = value.upper()'''
# Update the entire column in one request
'''sheet.updateColumn(1,columnOne)'''
# Getting all the Rows
#print(sheet.getRows())
'''rows = sheet.getRows()'''
# Gets the first row as a list in a variable
'''print(rows[1])
print(rows[10])'''

# Getting the items from the rows
'''print(rows[10][2])
print(rows[10][3])
'''
# Updating the list items in that row
'''rows[10][2] = '400'
rows[10][3] = '904'
print(rows[10])'''
# Updating the online spreadsheet with the changes
'''sheet.updateRows(rows)'''
# Getting the number of rows and columns in the sheet
'''print(sheet.rowCount)
print(sheet.columnCount)'''
# Changing the number of columns to 4
'''sheet.columnCount = 4
print(sheet.columnCount)'''

## Creating and Deleting Sheets

# Creating a new blank worksheet with only one sheet
#ss = ezsheets.createSpreadsheet('Multiple Sheets')
'''ss = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/1AgirLwhd0NBVYDxv9q8hVe4uHWoNnwcX4g6xxXOqVis/edit#gid=0')
print(ss.sheetTitles)'''
# Creating a new sheet at the end of the list of sheets
'''ss.createSheet('Spam')'''
# Create another sheet
'''ss.createSheet('Eggs')'''
# Check all the new sheets
'''print(ss.sheetTitles)'''
# Creating a new sheet at the begining of the list of sheets
'''ss.createSheet('Bacon',0)'''
# Check all the new sheets
'''print(ss.sheetTitles)'''
# Will delete the first item from the list of sheets
'''ss[0].delete()
print(ss.sheetTitles)'''
# You can also inditicate with the title name
'''ss['Spam'].delete()
print(ss.sheetTitles)'''
# Assign a variable to the eggs sheet and then delete it
'''sheet = ss['Eggs']
sheet.delete()
print(ss.sheetTitles)'''
# Adding a value in a cell in that sheet then clearing it but still keeping the sheet
'''sheet = ss[0]
sheet['A1'] = 'Hello There'
sheet.clear()
print(ss.sheetTitles)'''

## Copying Sheets

# Creating two new blank spreadsheets.
#ss1 = ezsheets.createSpreadsheet('First Spreadsheet')
'''ss1 = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/1BMs9ciSaVdJ40cPJb1SDJKdO4SXMRfaRqMh6A9h016Y/edit?ouid=116545985698011037990&usp=sheets_home&ths=true')'''
#ss2 = ezsheets.createSpreadsheet('Second Spreadsheet')
'''ss2 = ezsheets.Spreadsheet('https://docs.google.com/spreadsheets/d/1PJI2o84F0hvtkaq-JIb8deonYeHkSo9mqISRvFlFE28/edit?ouid=116545985698011037990&usp=sheets_home&ths=true')'''

# Getting a sheet from the spreadsheet
'''print(ss1[0])'''
# Entering the new data to the first row
'''ss1[0].updateRow(1,['Some','Data','in','First','Row'])
sheet = ss1[0]
sheet.columnCount = 5
sheet.rowCount = 2'''
# Copying sheet to another spreadsheet
# it will create a new sheet instead of overwriting to default ones
'''sheet.copyTo(ss2)'''
# Checking the sheets of ss2
'''print(ss2.sheetTitles)'''
# Changing the title of the copy of the sheet
'''ss2[0].title = 'Copy of SS1'
print(ss2.sheetTitles)
'''

## ------------------- Chapter 15 Working With PDF and Word Documents ---------------------##

# From the video
'''
import PyPDF2

import os 

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')

pdfFile = open('meetingminutes.pdf','rb')

reader = PyPDF2.PdfFileReader(pdfFile)

print(reader.numPages)

page = reader.getPage(0)

print(page.extractText())
'''
'''
for pageNum in range(reader.numPages):
    print(reader.getPage(pageNum).extractText())
'''
'''

pdf1File = open('meetingminutes.pdf','rb')

pdf2File = open('meetingminutes2.pdf','rb')

reader1 = PyPDF2.PdfFileReader(pdf1File)
reader2 = PyPDF2.PdfFileReader(pdf2File)

writer = PyPDF2.PdfFileWriter()

for pageNum in range(reader1.numPages):
    page = reader1.getPage(pageNum)
    writer.addPage(page)

for pageNum in range(reader2.numPages):
    page = reader2.getPage(pageNum)
    writer.addPage(page)

outputFile = open('combinedminutes.pdf','wb')
writer.write(outputFile)
outputFile.close()
pdf1File.close()
pdf2File.close()
'''

# From the book

## Extracting Text from PDF document.

# First import the module
'''import PyPDF2'''
# Then open the pdf file with 'rb' read binary mode and store it in a variable
'''pdfFileObj = open('meetingminutes.pdf','rb')
print(type(pdfFileObj))'''
# To get a PdfFileReader object that represents this PDF, call PyPDF2.PdfFileReader()
'''pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
print(type(pdfReader))'''
# Total number of Pages in that PDF document with numPages attribute
'''print(pdfReader.numPages)'''
# To extract a page from this PDF you need a Page object which you can get with getPage() method from PdfFileReader object.
'''pageObj = pdfReader.getPage(0)'''
# Once you have a page object you can call extractText() method to return a string of the page's text.
'''print(pageObj.extractText())
pdfFileObj.close()'''

# Decrypting PDF documents.
'''
import PyPDF2
import os
os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')
'''
# Getting a pdfReader obj
'''pdfReader = PyPDF2.PdfFileReader(open('encrypted.pdf','rb'))'''
# This will give you a True value which means it is password protected
'''print(pdfReader.isEncrypted)'''
# This will give you an error
'''pdfReader.getPage(0)'''
# We can decrypt the file with its password such as below:
'''print(pdfReader.decrypt('rosebud'))'''
# After decryption the file will give us a page obj like so:
'''pageObj = pdfReader.getPage(0)'''

# Creating PDF documents.

# Copying Pages

'''import PyPDF2
import os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')'''

# Opening 2 PDF files and getting PdfFile Object
'''pdf1File = open('meetingminutes.pdf','rb')
pdf2File = open('meetingminutes2.pdf','rb')'''
# Turning those PDFFile objects into a Reader object
'''pdf1Reader = PyPDF2.PdfFileReader(pdf1File)
pdf2Reader = PyPDF2.PdfFileReader(pdf2File)'''
# Creating a new Writer object
'''pdfWriter = PyPDF2.PdfFileWriter()'''

# Copying first file and getting a page object and copying that page object into a writer object
'''for pageNum in range(pdf1Reader.numPages):
    pageObj = pdf1Reader.getPage(pageNum)
    pdfWriter.addPage(pageObj)
'''
# Same as above but for the second file
'''for pageNum in range(pdf2Reader.numPages):
    pageObj = pdf2Reader.getPage(pageNum)
    pdfWriter.addPage(pageObj)
'''
# Getting a new pdf object in a write binary mode and copying the writer object into it
'''pdfOutputFile = open('combinedminutes2.pdf','wb')
pdfWriter.write(pdfOutputFile)
pdfOutputFile.close()
pdf1File.close()
pdf2File.close()'''

# Rotating Pages

'''import PyPDF2,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')'''

# Getting the File object
'''minutesFile = open('meetingminutes.pdf','rb')'''
# Getting the reader Object
'''pdfReader = PyPDF2.PdfFileReader(minutesFile)'''
# Getting the page object
'''page = pdfReader.getPage(0)'''
# Rotating that page 0
'''print(page.rotateClockwise(90))'''

# Creating a writer Object
'''pdfWriter = PyPDF2.PdfFileWriter()'''
# Adding that page object into a writer object
'''pdfWriter.addPage(page)'''
# Create a new blank pdf file in write-binary mode
'''resultPdfFile = open('rotatedPage.pdf','wb')'''
# Writing the writer object into a new blank file object
'''pdfWriter.write(resultPdfFile)
resultPdfFile.close()
minutesFile.close()'''

# Overlaying Pages

'''import PyPDF2,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')'''

# Getting the file object and reader object and page object from the source pdf file
'''minutesFile = open('meetingminutes.pdf','rb')
PdfReader = PyPDF2.PdfFileReader(minutesFile)
minutesFirstPage = PdfReader.getPage(0)'''

# Getting the watermark pdf file as a reader object 
'''pdfWatermarkReader = PyPDF2.PdfFileReader(open('watermark.pdf','rb'))'''
# merging the watermark reader object to the first page of source pdf page object
'''minutesFirstPage.mergePage(pdfWatermarkReader.getPage(0))'''

# Creating a new writer object and adding the watermarked firstpage into it.
'''pdfWriter = PyPDF2.PdfFileWriter()
pdfWriter.addPage(minutesFirstPage)'''

# Copying the source pdf pages and adding them to a writer object
'''for pageNum in range(1,PdfReader.numPages):
    pageObj = PdfReader.getPage(pageNum)
    pdfWriter.addPage(pageObj)'''

# Creating a new result pdf file object with write-binary mode and copying the contents
'''resultPdfFile = open('watermarkedCover.pdf','wb')
pdfWriter.write(resultPdfFile)
minutesFile.close()
resultPdfFile.close()'''

# Encrypting PDFs

'''import PyPDF2,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')'''

# opening the Source pdf File object reader object and Creating a writer object
'''pdfFile = open('meetingminutes.pdf','rb')
pdfReader = PyPDF2.PdfFileReader(pdfFile)
pdfWriter = PyPDF2.PdfFileWriter()
'''
# Copying the source pdf reader page objects into a writer object
'''for pageNum in range(pdfReader.numPages):
    pdfWriter.addPage(pdfReader.getPage(pageNum))
'''
# Adding encryption to a writer object, passing the password as a string
'''pdfWriter.encrypt('swordfish')'''

# Creating a new pdf file and copying the writer object contents into it
'''resultPdf = open('encryptedminutes.pdf','wb')
pdfWriter.write(resultPdf)
resultPdf.close()'''

## ------- Word Documents ------- ##

## From the video

'''import docx,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')
'''
'''d = docx.Document('demo.docx')

print(d.paragraphs)
print(type(d.paragraphs))
print(d.paragraphs[0])

print(d.paragraphs[0].text)

print(d.paragraphs[1].text)

p = d.paragraphs[1]
print(type(p))
print(p.runs)

print(p.runs[0].text),print(p.runs[2].text), print(p.runs[1].text)
p.runs[3].underline = True
p.runs[3].italic = True
p.runs[4].text = ''
p.runs[3].text = ' italic and underlined '

p.style = 'Title'
d.save('demo2.docx')
'''

'''d = docx.Document()
d.add_paragraph('Hello this is a paragraph')
d.add_paragraph('Hello this is another paragraph')
d.add_paragraph('Yet Hello again and this one is the final paragraph')

d.save('demo3.docx')

p = d.paragraphs[0]
p.add_run(' and this will be the paragraph\'s add run method')

print(p.runs)

p.runs[1].bold = True

d.save('demo3.docx')'''

'''def getText(filename):
    doc = docx.Document(filename)
    fullText = []
    for paragraph in doc.paragraphs:
        fullText.append(paragraph.text)
    return '\n'.join(fullText)

print(getText('demo.docx'))'''

## From the book 

# Reading Word Documents

# Creating a document object
'''doc = docx.Document('demo.docx')
print(type(doc))
'''
# Creating a list class para object with .paragraphs attribute
'''para = doc.paragraphs
print(type(para))'''
# There are 7 paragraph objects in that document
'''print(len(doc.paragraphs))'''
# Para objects has text attribute which contains a string in that paragraph without the style info
'''print(para[0].text)
print(para[1].text)'''
# Each para object also has a .runs attribute which is a list of run objects.
'''print(len(para[1].runs))'''
# Runs end with a style change
'''print(para[1].runs[0].text)
print(para[1].runs[1].text)
print(para[1].runs[2].text)
print(para[1].runs[3].text)
print(para[1].runs[4].text)'''

# Getting the Full Text from a .docx File

# This function opens a word document, loops over all the Paragraph objects in the paragraphs list, and then appends their text to the list fullText. 
# After the loop, the string in the fullText are joined by \n newline characters.

'''def getText(filename):
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
        # to indent each paragraph use this:
        fullText.append(' ' + para.text)
    return '\n'.join(fullText)
    # To add a double space between paragraphs use this:
    return '\n\n'.join(fullText)

print(getText('demo3.docx'))'''

# Run Attributes

'''doc = docx.Document('demo.docx')
print(doc.paragraphs[0].text)

print(doc.paragraphs[0].style)

doc.paragraphs[0].style = 'Normal'

print(doc.paragraphs[1].text)

doc.paragraphs[1].runs[0].style = 'QuoteChar'
doc.paragraphs[1].runs[1].underline = True
doc.paragraphs[1].runs[4].underline = True

doc.save('restyled.docx')'''

# Writing Word Documents

# Create a new blank Document object
'''doc = docx.Document()'''
# Use add_paragraph method of Document object
'''doc.add_paragraph('Hello Word!')'''
# Save the document
'''doc.save('helloword2.docx')'''
# Adding an additinonal paragraphs, Notice the second argument as a style
'''paraObj1 = doc.add_paragraph('This is a second paragraph','Title')
paraObj2 = doc.add_paragraph('This is yet another paragraph','Quote')'''
# Adding a run object to the paragraph, Notice the second argument as a style
'''paraObj1.add_run(' This text is being added to the second paragraph','QuoteChar')'''
# Save the document
'''doc.save('helloworld2.docx')'''

# Adding Headings

# Creating a new blank docx document
'''doc = docx.Document()'''
# add_heading() accepts a string value for the heading and an integer value from 0 to 4
# 0 is the Title style, the rest are various levels of headers from 1 being the biggest to smallest which is 4.

'''doc.add_heading('Header 0',0)

doc.add_heading('Header 1',1)

doc.add_heading('Header 2',2)

doc.add_heading('Header 3',3)

doc.add_heading('Header 4',4)

doc.save('headings1.docx')'''

# Adding Line and Page Breaks

# Creating the Document object
'''doc = docx.Document()'''
# Adding a paragraph Object
'''doc.add_paragraph('This is on the first page')'''
# Adding a pagebreak after the first paragraph
'''doc.paragraphs[0].runs[0].add_break(docx.enum.text.WD_BREAK.PAGE)'''
# Add another paragraph
'''doc.add_paragraph('This is on the second page')'''
# Add a line break after this run of that paragraph
'''doc.paragraphs[1].runs[0].add_break()'''
# This will be on the new line but still the same paragraph
'''doc.paragraphs[1].add_run(' Is this a good idea?')'''
# Save the Document
'''doc.save('twopage2.docx')'''

# Adding Pictures

# Creating a Document object
'''doc = docx.Document()'''
# Using document's add_picture() method, First is the string of the filename.
# Optional second and third are the diameters can be both inches or centimetres.
'''doc.add_picture('zophie.png',width=docx.shared.Inches(1),height=docx.shared.Cm(4))'''
# Save the document
'''doc.save('Zophie.docx')'''

# Creating PDFs from Word Documents

'''import win32com.client,docx,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')

wordFilename = 'your_word_document.docx'
pdfFilename = 'your_pdf_filename.pdf'

doc = docx.Document()
# Code to create Word document goes here.
doc.add_paragraph('This is a test of love innit?')

doc.save(wordFilename)

# Word's numeric code for PDFs.
wdFormatPDF = 17
wordObj = win32com.client.Dispatch('Word.Application')

docObj = wordObj.Documents.Open(wordFilename)
docObj.SaveAs(pdfFilename,FileFormat=wdFormatPDF)
docObj.Close()
wordObj.Quit()'''

## -------------- Chapter 16: Working With CSV files and JSON data -----------------------##

# The CSV module

# Reader Objects

'''import csv,os
os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')'''
# To read a CSV file, first open the file using the open() Function
'''exampleFile = open('example.csv')
print(type(exampleFile))'''
# Instead of passing the File object to a read() or readlines() function we passed it into a csv.reader() function therefore getting(creating) a Reader Object.
'''exampleReader = csv.reader(exampleFile)
print(type(exampleReader))
print(exampleReader)'''
# The most direct way of accessing the values in the Reader object is converting it into a plan list type with list() function.
'''exampleData = list(exampleReader)
print(exampleData)'''
# Now that we have a CSV file as a list of lists, we can access the value at a particular row and column with the expression exampleData[row][col] ==> where row is the index of one of the lists in exampleData and col is the index of the item you want from that list.
'''print(exampleData[0][0])
print(exampleData[0][1])
print(exampleData[0][2])
print(exampleData[1][2])
print(exampleData[6][1])'''

# Reading Data from reader Objects in a for Loop

# Opening the csv file and converting into a csv.reader object
'''exampleFile = open('example.csv')
exampleReader = csv.reader(exampleFile)'''
# Looping through the reader object. Each row is a list of values, with each value representing a cell
'''for row in exampleReader:'''
    # print() function call prints the number of the current row and the contents of the row.
'''print('Row # ' + str(exampleReader.line_num)+ ' ' + str(row))'''
    # To get the row number, user the reader object's line_num variable

# Writer Objects

# Opening a new csv file in write mode. On windows you'll need to pass a blank string for the open() function's newline keyword.If not output.csv will be double-spaced.
'''outputFile = open('output.csv','w',newline='')'''
# Create a writer object using the file object
'''outputWriter = csv.writer(outputFile)'''
# Writing rows to that writer object using writerow()
'''outputWriter.writerow(['spam','eggs','ham'])'''
# Each line will be a new row in output.csv file.
'''outputWriter.writerow(['Hello, World!','eggs','bacon','ham'])
outputWriter.writerow([1,2,3.141592,4])'''
# Close the file object
'''outputFile.close()'''

# The Delimiter and lineterminator keyword Arguments
# Opening a example.tsv file in a write mode / we are using tsv because it is tab seperated and not comma separeted
'''csvFile = open('example.tsv','w',newline='')'''
# Delimiter will be a Tab Character and lineterminator will be doublespace
'''csvWriter = csv.writer(csvFile,delimiter='\t',lineterminator='\n\n')'''
# Writing 3 rows to a file and closing it.
'''csvWriter.writerow(['apples','oranges','grapes'])
csvWriter.writerow(['eggs','bacon','ham'])
csvWriter.writerow(['spam','spam','spam','spam','spam','spam'])
csvFile.close()'''

# DictReader and DictWriter CSV Objects

# Opening a new csv file with header and creating a DictReader object
'''exampleFile = open('exampleWithHeader.csv')
exampleDictReader = csv.DictReader(exampleFile)'''
# type of DictReader is Dictreader
'''print(exampleDictReader)
print(type(exampleDictReader))'''
# For every row in that file, headers are stored as keys and values are stored inside of it.
'''for row in exampleDictReader:
    print(row['Timestamp'],row['Fruit'],row['Quantity'])'''

# Creating our own headers in DictReader objects

# opening csv file withouth the header
'''exampleFile1 = open('example.csv')'''
# Adding out own made up headers as a secon argument here
'''exampleDictReader1 = csv.DictReader(exampleFile1,['time','name','amount'])'''
# printing the values of rows just the same
'''for row in exampleDictReader1:
    print(row['time'],row['name'],row['amount'])'''

# DictWriter Object

# Opening a new csv file in write mode
'''outputFile = open('output.csv','w',newline='')'''
# Creating a DictWriter object and passing our made up headers as a second argument
'''outputDictWriter = csv.DictWriter(outputFile,['Name','Pet','Phone'])'''
# saving our header names in that DictWriter object
'''outputDictWriter.writeheader()'''
# Writing additional rows to the file notice the order doesn't matter
'''outputDictWriter.writerow({'Name':'Alice','Pet':'cat','Phone':'555-1234'})
outputDictWriter.writerow({'Name':'Bob','Phone':'555-9999'})
outputDictWriter.writerow({'Phone':'555-5555','Name':'Carol','Pet':'dog'})'''
# Closing the file
'''outputFile.close()'''

## Project: Removing the Header From CSV Files ## 

# Program will do the following:
    # 1. Find all the CSV files in the current working directory.
    # 2. Read in the full contents of each file.
    # 3. Write out the contents, skipping the first line, to a new CSV file.

# At the code level program will do the following:
    # 1. Loop over a list of files from os.listdir(), skipping the non-CSV files.
    # 2. Create a CSV reader object and read in the contents of the file, using the line_num attribute to figure out which line to skip.
    # 3. Create a CSV writer object and write out the read-in data to the new file.

## STEP 1: Loop through each CSV file
## STEP 2: Read in the CSV file
## STEP 3: Write out the CSV file Without the first row
'''import csv,os

os.chdir('C:\\Users\\devil\\Desktop\\Python Beginners Code and Basics\\Automate the boring staff with Python\Files used in the book\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')

os.makedirs('headerRemoved',exist_ok=True)

# Loop through every file in the current working directory.

for csvFilename in os.listdir('.'):
    if not csvFilename.endswith('.csv'):
        continue                            # Skip the non-csv files
    print('Removing header from ' + csvFilename + '...')

    # TODO : Read the CSV file in (Skipping first row).
    csvRows = []
    csvFileObj = open(csvFilename)
    readerObj = csv.reader(csvFileObj)
    for row in readerObj:
        if readerObj.line_num == 1:
            continue                        # Skip the first row
        csvRows.append(row)
    csvFileObj.close()
    # TODO : Write out the CSV file.
    csvFileObj = open(os.path.join('headerRemoved',csvFilename),'w',newline='')
    csvWriter = csv.writer(csvFileObj)
    for row in csvRows:
        csvWriter.writerow(row)
    csvFileObj.close()'''

## Ideas for Similar programs:

    # Compare data between different rows in a CSV file or between multiple CSV files.
    # Copy specific data from CSV file to an Excel file, or vice versa.
    # Check for invalid data or formatting mistakes in CSV files and alert the user to these errors.
    # Read data from a CSV file as input for your Python programs.

## JSON and APIs ## 

# Reading JSON with the loads() Function.

# Notice the double quotes, JSON strings always use double-quotes.
'''stringOfJsonData = '{"name":"Zophie","isCat":true,"miceCought":0,"felineIQ":null}'
print(type(stringOfJsonData))
import json'''
# passing the json data into a json.loads() and returning a Python value.
'''jsonDataAsPythonValue = json.loads(stringOfJsonData)'''
# See the result Python Value which is a dictionary.
'''print(jsonDataAsPythonValue)
print(type(jsonDataAsPythonValue))'''

# Writing JSON with the dumps() Function

# Getting the dict data type value
'''pythonValue = {'isCat':True,'miceCought':0,'name':'Zophie','felineIQ':None}
print(type(pythonValue))'''
# import json module
'''import json'''
# Converting python value into a string of JSON formatted data
'''stringOfJsondata = json.dumps(pythonValue)'''
# Result string and data type
'''print(stringOfJsondata)
print(type(stringOfJsondata))'''

## PROJECT: FETCHING CURRENT WEATHER DATA ##


# This program does the following:
    # 1. Reads the requested location from the command line
    # 2. Downloads JSON weather data from OpenWeatherMap.org
    # 3. Converts the string of JSON data to a Python data structure.
    # 4. Prints the wather for today and th next two days.

# Code will need to do the following:
    # 1. Join strings in sys.argv to get the location.
    # 2. Call requests.get() to download the weather data.
    # 3. Call json.loads() to convert the JSON data to a Python data structure.
    # 4. Print the weather forecast.


# STEP 1 : Get Location from the Command line Argument

'''
#! python3
#getOpenWeather.py -- Prints the weather for a location from the command line

APPID = '568e213d7755eb31f25278300108aa8f'

import json,requests,sys

# Compute location from comman line arguments.

if len(sys.argv) < 2:
    print('Usage: getOpenWeather.py city_name, 2-letter_country_code')
    sys.exit()

location = ' '.join(sys.argv[1:])

# TODO : Download the JSON data from OpenWeatherMap.org's API

# TODO : Load JSON data into a Python Variable

'''

# STEP 2 : Download the JSON Data.

'''
# TODO : Download the JSON data from OpenWeatherMap.org's API

url = 'https://api.openweathermap.org/data/2.5/forecast/daily?q=%s&cnt=3&APPID=%s'% (location,APPID)
response = requests.get(url)
response.raise_for_status()
# Uncomment to see the raw JSON text:
#print(response.text)
'''

# STEP 3 : Load JSON Data and Print Weather

'''
# Print Weather Descriptions.

w = weatherData['list']
print('Current weather in %s:'% (location))
print(w[0]['weather'][0]['main'],'-',w[0]['weather'][0]['description'])
print()
print('Tomorrow:')
print(w[1]['weather'][0]['main'],'-',w[1]['weather'][0]['description'])
print()
print('Day after Tomorrow:')
print(w[2]['weather'][0]['main'],'-',w[2]['weather'][0]['description'])

'''

# Ideas For Similar Progams:
    # 1. Collect weather Forecasts for several campsites or hiking trails to see which one will have the best weather.
    # 2. Schedule a program to regularly check the weather and send you a frost alert if you need to move your plant indoors.(Chapter 17 covers schedulling and Chapter 18 explains how to send email.)
    # 3. Pull weather data from multiple sites to show all at once, or calculate and show the average of the multiple weather predictions.


## -------- Chapter 17 : Keeping Time,Scheduling Tasks,And Launching Programs -------------##

# The Time Module

# the time.time() Function

'''import time

print(time.time())'''

# We define a functionto loop throughthe integers from 1 to 99,999 and return product.
'''def calcProd():
    # Calculate the product of the first 100,000 numbers.
    product = 1
    for i in range(1,100000):
        product = product*i
    return product'''
# We call time.time() function at the start and store it in variable called startTime
'''startTime = time.time()'''
# Calling the function
'''prod = calcProd()'''
# We call time.time() function again at the end of the function and store the value in endTime variable
'''endTime = time.time()'''
# We are printing the length of the result
'''print('The results is %s digits long'%(len(str(prod))))'''
# We also print the time it took to calculate with substracting the two values.
'''print('Took %s seconds to calculate'% (endTime - startTime))'''

# For more info about cProfile.run() Function visit: https://docs.python.org/3/library/profile.html

'''import time
'''
# Will return human readable time in string format
'''print(time.ctime())
print(type(time.ctime()))
'''
# Can also use to convert Unix epoch into a string value
'''thisMoment = time.time()
print(time.ctime(thisMoment))'''

# The time.sleep() Function

'''import time
'''
# The for loop will print 'Tick' and pause for 1 second then print 'Tock' and pause for another 1 second, until Tick and Tock have been printed three times.
'''for i in range(3):
    print('Tick')
    time.sleep(1)
    print('Tock')
    time.sleep(1)
'''
# Rounding Numbers

'''import time

now = time.time()
'''
# Original number
'''print(now)'''
# Round two digits after the decimal
'''print(round(now,2))'''
# Round four digits after the decimal
'''print(round(now,4))'''
# Round to the nearest integer.
'''print(round(now))'''

# PROJECT: Super Stopwatch

# Your program will do :
    # 1. Track the amount of time elapsed between presses of the ENTER key, with each key press starting a new 'lap' on the timer.
    # 2. Print the lap number,total-time, and lap time.

# This means your code will do the following:
    # 1. Find the current time by calling time.time() function and store it as a timestamp at the start of the program, as well as at the start of each lap.
    # 2. Keep a lap counter and increment it every time the user presses ENTER.
    # 3. Calculate the elapsed time by substracting timestamps.
    # 4. Handle the KeyboardInterrupt exception so the user can press Ctrl+C to quit.

# STEP 1 : Set up the program to track times.

    # The stopwatch program will need to use the current time, so you’ll want to import the time module. Your program should also print some brief instructions to the user before calling input(), so the timer can begin after the user presses ENTER. Then the code will start tracking lap times.

'''
#! python3
# stopwatch.py -- Simple stopwatch program

import time

# Display the program's instructions.

print('Press ENTER to begin. Afterward, press ENTER to "click" the stopwatch.\nPress Ctrl+C to quit.')

# Press ENTER to begin
input()

print('Started')

# Get the first lap's start time.
starTime = time.time()
lastTime = starTime

lapNum = 1

# TODO : Start tracking the lap times
'''

# STEP 2 : Track and print Lap times.

    # Now let’s write the code to start each new lap, calculate how long the previous lap took, and calculate the total time elapsed since starting the stopwatch. We’ll display the lap time and total time and increase the lap count for each new lap
'''
try:
    while True:
        input()
        lapTime = round(time.time() - lastTime,2)
        totalTime = round(time.time() - starTime,2)
        print('Lap #%s: %s(%s)'%(lapNum,totalTime,lapTime),end='')
        lapNum += 1
        # Reset the last lap time
        lastTime = time.time()
except KeyboardInterrupt:
    # Handle the Ctrl+C exception to keep its error message from displaying.
    print('\nDone...')
    
'''
# If the user presses CTRL-C to stop the stopwatch, the KeyboardInterrupt exception will be raised, and the program will crash if its execution is not a try statement. To prevent crashing, we wrap this part of the program in a try statement ➊. We’ll handle the exception in the except clause ➏, so when CTRL-C is pressed and the exception is raised, the program execution moves to the except clause to print Done, instead of the KeyboardInterrupt error message. Until this happens, the execution is inside an infinite loop ➋ that calls input() and waits until the user presses ENTER to end a lap. When a lap ends, we calculate how long the lap took by subtracting the start time of the lap, lastTime, from the current time, time.time() ➌. We calculate the total time elapsed by subtracting the overall start time of the stopwatch, startTime, from the current time ➍.

# Since the results of these time calculations will have many digits after the decimal point (such as 4.766272783279419), we use the round() function to round the float value to two digits at ➌ and ➍.

# At ➎, we print the lap number, total time elapsed, and the lap time. Since the user pressing ENTER for the input() call will print a newline to the screen, pass end='' to the print() function to avoid double-spacing the output. After printing the lap information, we get ready for the next lap by adding 1 to the count lapNum and setting lastTime to the current time, which is the start time of the next lap.

## Ideas for Similar Programs:

    # Create a simple timesheet app that records when you type a person's name and uses current time to clock them in or out.
    # Add a feature to your program to display the elapsed time since a process started, such as a download that uses the requests module.
    # Intermittently check how long a program has been running and offer the user a chance to cancel tasks that are taking too long.

## The Datetime Module

'''import datetime'''

# datetime.datetime.now() returns a datetime object for the current date and time
# This object includes year,month,day,hour,minute,second and milisecond of the current moment
'''print(datetime.datetime.now())
print(type(datetime.datetime.now()))'''
# We can also retrieve a datetime object by using datetime.datetime() function.
# Passing the integers representing year,month,day,hour,minute,second of the moment you want.
'''dt = datetime.datetime(2019,10,21,16,29,0)'''
# These integers will be stored in the datetime object's year,month,day,hour,minute and second attributes
'''print(dt.year,dt.month,dt.day)
print(dt.hour,dt.minute,dt.second)'''

'''import datetime,time
'''
# Calling datetime.datetime.fromtimestamp(1000000) and passing it 1000000 returns a datetime object for the moment 1000000 seconds after Unix epoch
'''print(datetime.datetime.fromtimestamp(1000000))'''
# datetime.datetime.now() and below are the same
'''print(datetime.datetime.fromtimestamp(time.time()))'''

# Comparing datetime objects

'''halloween2019 = datetime.datetime(2019,10,31,0,0,0)
newyears2020 = datetime.datetime(2020,1,1,0,0,0)
oct31_2019 = datetime.datetime(2019,10,31,0,0,0)'''

'''print(halloween2019 == oct31_2019)
print(halloween2019 > newyears2020)
print(newyears2020 > halloween2019)
print(newyears2020 != oct31_2019)'''

# The timedelta Data type

'''import datetime'''

# We create a timedelta object as below. Specifying the duration of days,hours,minutes and seconds.
'''delta = datetime.timedelta(days=11,hours=10,minutes=9,seconds=8)
print(delta)
print(type(delta))'''
# This object holds days attribute in days which is 11, and its seconds attribute stores 36548(10 hours,9 minutes,8 seconds)
'''print(delta.days,delta.seconds,delta.microseconds)'''
# Total seconds value of 11 days 10 hours 9 minutes 8 seconds
'''print(delta.total_seconds())'''
# Passing into a str() function to return a human-readable value.
'''print(str(delta))'''

# Arithmetic operators

'''import datetime'''

# To calculate the date 1,000 days from now;

# First we make a datetime object for the current moment and store it in 'dt'
'''dt = datetime.datetime.now()
print(dt)'''
# Then we make timedelta object for duration 1000 days and store it in 'thousandDays'
'''thousandDays = datetime.timedelta(days=1000)'''
# Finally we add both of the objects and get the return value.
'''print(dt+thousandDays)'''
# Creating a timedelta object for thirty years
'''aboutThirtyYears = datetime.timedelta(days=365*30)'''
# Adding and substracting current moment with thirty years.
'''print(dt+aboutThirtyYears)
print(dt-aboutThirtyYears)'''

# Pausing until a Specific Date

# time.sleep() method lets you pause a program for a certain number of seconds.By using a While loop, you can pause your programs until a specific date.

'''import datetime,time

halloween2016 = datetime.datetime(2021,8,13,22,20,0)

while datetime.datetime.now() < halloween2016:
    time.sleep(1)
    print('It\' time to shineee')'''

## Converting datetime Objects into String 

'''import datetime

dt = datetime.datetime.now()

print(dt.strftime('%Y/%m/%d %H:%M:%S'))

print(dt.strftime('%I:%M %p'))

print(dt.strftime("%B of '%y'"))'''

# Converting String into datetime Objects

'''import datetime
from time import strptime
'''
# To get a datetime object pass that string as the first argument and the custom format string as the second argument.
'''print(datetime.datetime.strptime('October 21, 2019', '%B %d, %Y'))
print(datetime.datetime.strptime('2019/10/21 16:29:00', '%Y/%m/%d %H:%M:%S'))
print(datetime.datetime.strptime("october of '19'","%B of '%y'"))
print(datetime.datetime.strptime("November of '63'","%B of '%y'"))'''
# The string with the date information must match the customformat string exactly or Python will raise a ValueError exception.

## MULTITHREADING ##

'''import threading,time'''

# Creating and testing the threading module

'''print('Start of Program.')'''
# Defining a function we want to use in a new thread
'''def takeANap():
    time.sleep(5)
    print('Wake up!!')'''
# Creating a thread Object ==> Notice its not target=takeANap()
'''threadObj = threading.Thread(target=takeANap)'''
# We are starting the executing the new thread
'''threadObj.start()'''
# This will be the original thread
'''print('End of Program.')
'''
# Normally a program terminates when the last line of code in the file has run (or the sys.exit() function is called). But the code above has two threads, Therefore it terminates when both of its threads terminated.

# Passing Arguments to the Thread's Target Function

'''print('Cats','Dogs','Frogs',sep='&')

import threading
'''
# We are calling the function's arguments in to args= keyword and keyword arguments(sep=) in to kwargs= keyword argument.As a list and dictionary respectively.
'''threadObj = threading.Thread(target=print,args=['Cats','Dogs','Frogs'],kwargs={'sep':'&'})'''
# Starting the new thread
'''threadObj.start()
'''
## Incorect way of doing it ==> 
'''
threadObj = threding.Thread(target=print('Cats','Dogs','Frogs',sep='%'))
Above will pass the function's return value as the target= keyword argument in 'print()' function's case it is None value
'''

# PROJECT : MULTITHREADED XKCD DOWNLOADER

## In Chapter 12, you wrote a program that downloaded all of the XKCD comic strips from the XKCD website. This was a single-threaded program: it downloaded one comic at a time. Much of the program’s running time was spent establishing the network connection to begin the download and writing the downloaded images to the hard drive. If you have a broadband internet connection, your single-threaded program wasn’t fully utilizing the available bandwidth.

## A multithreaded program that has some threads downloading comics while others are establishing connections and writing the comic image files to disk uses your internet connection more efficiently and downloads the collection of comics more quickly. Open a new file editor tab and save it as threadedDownloadXkcd.py. You will modify this program to add multithreading. The completely modified source code is available to download from https://nostarch.com/automatestuff2/.

# STEP 1 : Modify the Program to use a Function

## This program will mostly be the same downloading code from Chapter 12, so I’ll skip the explanation for the requests and Beautiful Soup code. The main changes you need to make are importing the threading module and making a downloadXkcd() function, which takes starting and ending comic numbers as parameters.

## For example, calling downloadXkcd(140, 280) would loop over the downloading code to download the comics at https://xkcd.com/140/, https://xkcd.com/141/, https://xkcd.com/142/, and so on, up to https://xkcd.com/279/. Each thread that you create will call downloadXkcd() and pass a different range of comics to download.

'''
import requests,os,bs4,threading

# Store comics in ./xkcd
os.makedirs('xkcd',exist_ok=True)

def downloadXkcd(startComic,endComic):
    for urlNumber in range(startComic,endComic):
        # Download the page.
        print('Downloading page https://xkcd.com/%s...'%(urlNumber))
        res = requests.get('https://xkcd.com/%s'%(urlNumber))
        res.raise_for_status()

        soup = bs4.BeautifulSoup(res.text,'lxml')

        # Find the URL of the comic image.
        comicElem = soup.select('#comic img')
        if comicElem == []:
            print('Could not find comic image')
        else:
            comicUrl = comicElem[0].get('src')
            # Download the image
            print('Downloading image %s...'%(comicUrl))
            res = requests.get('https://'+comicUrl)
            res.raise_for_status()

            # Save the image to ./xcd.
            imageFile = open(os.path.join('xkcd',os.path.basename(comicUrl)),'wb')
            for chunk in res.iter_content(100000):
                imageFile.write(chunk)
            imageFile.close()

# TODO : Create and start the Thread Objects.

# TODO : Wait for all threads to end.

After importing the modules we need, we make a directory to store comics in ➊ and start defining downloadxkcd() ➋. We loop through all the numbers in the specified range ➌ and download each page ➍. We use Beautiful Soup to look through the HTML of each page ➎ and find the comic image ➏. If no comic image is found on a page, we print a message. Otherwise, we get the URL of the image ➐ and download the image ➑. Finally, we save the image to the directory we created.
'''
# STEP 2 : Create and Start Threads

# Now that we’ve defined downloadXkcd(), we’ll create the multiple threads that each call downloadXkcd() to download different ranges of comics from the XKCD website. 

'''
# TODO : Create and start the Thread Objects.
# a list of all he Thread Objects
downloadThreads = []
# Loops 14 times, creates 14 threads.
for i in range(0,140,10):
    start = i
    end = i+9
    if start == 0:
        # There is no comic 0, so set it to 1.
        start = 1
    downloadThread = threading.Thread(target=downloadXkcd,args=(start,end))
    downloadThreads.append(downloadThread)
    downloadThread.start()

First we make an empy list downloadThreads; the list will help us keep track of the many Thread objects we’ll create. Then we start our for loop. Each time through the loop, we create a Thread object with threading.Thread(), append the Thread object to the list, and call start() to start running downloadXkcd() in the new thread. Since the for loop sets the i variable from 0 to 140 at steps of 10, i will be set to 0 on the first iteration, 10 on the second iteration, 20 on the third, and so on. Since we pass args=(start, end) to threading.Thread(), the two arguments passed to downloadXkcd() will be 1 and 9 on the first iteration, 10 and 19 on the second iteration, 20 and 29 on the third, and so on.

As the Thread object’s start() method is called and the new thread begins to run the code inside downloadXkcd(), the main thread will continue to the next iteration of the for loop and create the next thread.

'''

# STEP 3 : Wait for All Threads to End.

# The main thread moves on as normal while the other threads we create download comics. But say there’s some code you don’t want to run in the main thread until all the threads have completed. Calling a Thread object’s join() method will block until that thread has finished. By using a for loop to iterate over all the Thread objects in the downloadThreads list, the main thread can call the join() method on each of the other threads

'''
# TODO : Wait for all threads to end.

for downloadThread in downloadThreads:
    downloadThread.join()
print('Done.')

The 'Done.' string will not be printed until all of the join() calls have returned. If a Thread object has already completed when its join() method is called, then the method will simply return immediately. If you wanted to extend this program with code that runs only after all of the comics downloaded, you could replace the print('Done.') line with your new code.
'''

## Launching other Programs from Python

# Creating a Popen object, passing the filename of calculator exe into it.
'''import subprocess
subprocess.Popen('C:\\Windows\\System32\\calc.exe')'''

# Here we open MS Paint process
'''paintProc = subprocess.Popen('C:\\Windows\\System32\\mspaint.exe')'''
# While it's still running we check wheter poll() returns None
'''print(paintProc.poll())'''
# It should as the process is  still running
'''if paintProc.poll() == None:
    print('It is none')'''
# Doesn't return until MS Paint closes.
'''print(paintProc.wait())'''
# Now wait() and poll() returns 0, indicating that the process terminated without errors
'''print(paintProc.poll())'''

# Passing Command Line Arguments to the Popen() Function.

# This process will both open notepad application but also immeddiatly open 'Programming.txt' file.
'''import subprocess
subprocess.Popen(['C:\\Windows\\notepad.exe','C:\\Users\\devil\\Programming.txt'])'''



# Running Other Python Scripts

'''import subprocess'''

# Popen() a list containing the Python.exe and a string of scripts's filename
'''print(subprocess.Popen(['C:\\Users\\devil\\AppData\\Local\\Programs\\Python\\Python39\\python.exe','C:\\Users\\devil\Desktop\\Python Beginners Code and Basics\\My Python Documents\\Census2010.py']))'''

# Opening Files with Default Applications

'''import subprocess'''
# Creating a new .txt file in write mode at cwd
'''fileObj = open('hello.txt','w')'''
# Writing a string to that .txt file
'''fileObj.write('Hello World!!')'''
# Closing the file
'''fileObj.close()'''
# Opening the hello.txt file with default(Notepad) Application using the 'start' keyword, shell=True is needed only on Windows
'''subprocess.Popen(['start','hello.txt'],shell=True)
'''
## PROJECT: Simple CountDown Program

# This program will do the following:
    # 1. Count down from 60.
    # 2. Play a sound file (alarm.wav) when the countdown reaches zero.

# Code will do the following:
    # 1. Pause for 1 second in between displaying each number in the countdown by calling time.sleep().
    # 2. Call subprocess.Popen() to open the sound file with the default application.

# STEP 1 : Count Down

'''
import subprocess,time

timeLeft = 60
while timeLeft > 0:
    print(timeLeft,end='')
    time.sleep(1)
    timeLeft = timeLeft - 1 

# TODO : At the end of the countdown, play a sound file.

After importing time and subprocess, make a variable called timeLeft to hold the number of seconds left in the countdown ➊. It can start at 60—or you can change the value here to whatever you need, or even have it get set from a command line argument.

In a while loop, you display the remaining count ➋, pause for 1 second ➌, and then decrement the timeLeft variable ➍ before the loop starts over again. The loop will keep looping as long as timeLeft is greater than 0. After that, the countdown will be over.
'''

# STEP 2 : Play the sound File.

'''

# TODO : At the end of the countdown, play a sound file.

subprocess.Popen(['start','alarm.wav'],shell=True)

After the while loop finishes, alarm.wav (or the sound file you choose) will play to notify the user that the countdown is over. On Windows, be sure to include 'start' in the list you pass to Popen() and pass the keyword argument shell=True. On macOS, pass 'open' instead of 'start' and remove shell=True.

Instead of playing a sound file, you could save a text file somewhere with a message like Break time is over! and use Popen() to open it at the end of the countdown. This will effectively create a pop-up window with a message. Or you could use the webbrowser.open() function to open a specific website at the end of the countdown. Unlike some free countdown application you’d find online, your own countdown program’s alarm can be anything you want!
'''

# Ideas for Similar Programs

# Use time.sleep() to give the user a chance to press CTRL-C to cancel an action, such as deleting files. Your program can print a "Press CTRL-C to cancel" message and then handle any KeyboardInterrupt exceptions with try and except statements.

# For a long-term countdown, you can use timedelta objects to measure the number of days,hours,minutes, and seconds until some point (a birthday? and anniversary?) in the future.

##------------------- Chapter 18: Sending Emails and Text Messages ------------------------##

